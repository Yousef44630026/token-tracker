"""Excel export via openpyxl — materializes the derived columns. (Phase 9)

One workbook, all sheets derived from the Trace (nothing stored):
  - TokenQuantities : the quantity-grain rows (quantity_in_total / export_warning);
  - TokenEvents     : the event-grain rows (event_contributing_tokens, 0 if superseded);
  - TokenSpans      : source-of-truth span identity and RAG/agent/tool metadata;
  - CoverageExactness + the other metric sheets (latency, reliability, cache, RAG, agent,
    service attribution), one per entry of ``build_metric_exports`` — the same shared table
    the CSV export writes, so both exports carry identical columns and values.
    CoverageExactness's observed total equals the model trace total by construction.
"""

from __future__ import annotations

from openpyxl import Workbook

from tracker.export.csv_exporter import (
    EVENT_HEADERS,
    METRIC_HEADERS,
    QUANTITY_HEADERS,
    SERVICE_ATTRIBUTION_HEADERS,
    SPAN_HEADERS,
    build_metric_exports,
    event_rows,
    quantity_rows,
    span_rows,
)
from tracker.models.trace import Trace


def _write_sheet(ws, headers, rows) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])


def export_excel(
    trace: Trace,
    path: str,
) -> str:
    """Write the four-sheet workbook to ``path`` and return it."""
    wb = Workbook()

    ws_q = wb.active
    ws_q.title = "TokenQuantities"
    _write_sheet(ws_q, QUANTITY_HEADERS, quantity_rows(trace))

    ws_e = wb.create_sheet("TokenEvents")
    _write_sheet(ws_e, EVENT_HEADERS, event_rows(trace))

    ws_s = wb.create_sheet("TokenSpans")
    _write_sheet(ws_s, SPAN_HEADERS, span_rows(trace))

    # CoverageExactness arrives FIRST from build_metric_exports (same sheet name/position as
    # before), so CSV and Excel are built from the one shared metric-export table.
    for sheet_name, rows in build_metric_exports(trace).items():
        ws = wb.create_sheet(sheet_name)
        headers = SERVICE_ATTRIBUTION_HEADERS if sheet_name == "ServiceAttribution" else METRIC_HEADERS
        _write_sheet(ws, headers, rows)

    wb.save(path)
    return path
