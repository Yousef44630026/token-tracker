"""Generate an audit-friendly native Excel dashboard from TokenEvent JSONL files.

Pricing is deliberately presentation-layer input. It is never written back to TokenEvent
or JSONL storage, preserving the tracker's source-vs-derived boundary.
"""

from __future__ import annotations

import argparse
import datetime as dt
import gzip
import json
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from tracker.analytics.provider_validation import (
    build_capability_certification_matrix,
    build_provider_validation_matrix,
    summarize_capability_certification,
)
from tracker.derive.effective_events import effective_events
from tracker.models.enums import DataQualityFlag, TokenType, Trust
from tracker.models.token_event import TokenEvent
from tracker.storage._locking import lock_for
from tracker.validation.fixture_manifest import PROVIDER_CAPABILITY_POLICIES, realistic_fixture_records

LOGGER = logging.getLogger("tracker.reporting.excel_dashboard")

PRICE_COLUMNS = (
    "provider",
    "model",
    "token_type",
    "price_per_million_tokens",
    "currency",
    "effective_from",
    "effective_to",
)

DATA_COLUMNS = (
    "source_file",
    "source_line",
    "source_kind",
    "source_event_id",
    "event_id",
    "request_correlation_id",
    "trace_id",
    "span_id",
    "parent_span_id",
    "business_id",
    "use_case",
    "workflow",
    "environment",
    "timestamp_utc",
    "date",
    "provider",
    "model",
    "deployment",
    "cloud_provider",
    "region",
    "service_name",
    "tenant_id",
    "provider_request_id",
    "provider_response_id",
    "retry_count",
    "api_surface",
    "event_authoritative",
    "event_status",
    "http_status",
    "event_superseded",
    "superseded_by",
    "data_quality_flags",
    "token_type",
    "raw_tokens",
    "precision_level",
    "usage_source",
    "overlap",
    "trust",
    "subtotal_of",
    "quantity_in_total",
    "exact_tokens_active",
    "estimated_tokens_active",
    "unverified_tokens_active",
    "event_contributing_tokens_once",
    "event_count_once",
    "event_authoritative_once",
    "quality_flagged_event_once",
    "active_quality_flagged_event_once",
    "superseded_event_once",
    "mismatch_event_once",
    "provider_total_tokens_once",
    "provider_total_observation_once",
    "event_total_mismatch_once",
    "under_attributed_tokens_once",
    "over_attributed_tokens_once",
    "schema_drift_event_once",
    "correlation_risk_event_once",
    "usage_loss_event_once",
    "billing_tokens",
    "billable_tokens_for_coverage",
    "priced_billing_tokens",
    "cache_read_tokens_active",
    "unknown_quantity_active",
    "unit_price_per_million",
    "currency",
    "derived_cost",
    "cost_quality",
    "request_count_once",
    "request_latency_applicable_once",
    "request_latency_ms",
    "request_latency_observation_once",
    "request_ttft_applicable_once",
    "request_ttft_ms",
    "request_ttft_observation_once",
    "quantity_metadata_json",
    "observation_json",
)

NAVY = "1F2937"
TEAL = "0F766E"
BLUE = "2563EB"
CORAL = "C2413B"
GOLD = "B7791F"
LIGHT = "F3F4F6"
WHITE = "FFFFFF"
MUTED = "6B7280"
GRID = "D1D5DB"
AZURE_BLUE = "0078D4"
FOUNDRY_TEAL = "008272"
SUCCESS = "107C10"
WARNING = "D83B01"
SOFT_BLUE = "E8F3FC"
MAX_DASHBOARD_DATES = 730
MAX_DASHBOARD_MODELS = 20
EXCEL_MAX_ROWS = 1_048_576
DEFAULT_MAX_DATA_ROWS = 250_000
DASHBOARD_ROW_WARNING = 100_000

USAGE_LOSS_FLAGS = frozenset(
    {
        DataQualityFlag.RAW_USAGE_MISSING.value,
        DataQualityFlag.PROVIDER_USAGE_MISSING.value,
        DataQualityFlag.PROVIDER_STREAM_USAGE_MISSING.value,
        DataQualityFlag.PROVIDER_RESPONSE_UNPARSEABLE.value,
        DataQualityFlag.NORMALIZATION_ERROR.value,
    }
)

DASHBOARD_FILTERS = (
    ("Provider", "provider", "B5", "DashboardProviders"),
    ("Model", "model", "D5", "DashboardModels"),
    ("Deployment", "deployment", "F5", "DashboardDeployments"),
    ("Environment", "environment", "H5", "DashboardEnvironments"),
    ("Use case", "use_case", "J5", "DashboardUseCases"),
)
DASHBOARD_FILTER_REFS = {
    column: f"'Dashboard'!${cell[0]}${cell[1:]}" for _, column, cell, _ in DASHBOARD_FILTERS
}
DASHBOARD_DATE_FROM = "'Dashboard'!$L$5"
DASHBOARD_DATE_TO = "'Dashboard'!$N$5"


@dataclass(frozen=True)
class LoadReport:
    files_read: int = 0
    lines_read: int = 0
    valid_events: int = 0
    malformed_lines: int = 0
    schema_invalid_lines: int = 0
    duplicate_event_ids: int = 0


@dataclass(frozen=True)
class LoadedEvent:
    event: TokenEvent
    source_file: str
    source_line: int
    sequence: int


def _parse_timestamp(value: str | None) -> dt.datetime | None:
    if not value:
        return None
    try:
        parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=dt.UTC)
    return parsed.astimezone(dt.UTC)


def _newer(candidate: LoadedEvent, current: LoadedEvent) -> bool:
    candidate_time = _parse_timestamp(candidate.event.timestamp)
    current_time = _parse_timestamp(current.event.timestamp)
    if candidate_time is not None and current_time is not None and candidate_time != current_time:
        return candidate_time > current_time
    if candidate_time is not None and current_time is None:
        return True
    if candidate_time is None and current_time is not None:
        return False
    return candidate.sequence > current.sequence


def load_jsonl_events(
    data_dir: str | os.PathLike[str],
    *,
    recursive: bool = False,
) -> tuple[list[LoadedEvent], LoadReport]:
    """Read and validate every JSONL row, logging malformed content without exposing it."""
    root = Path(data_dir)
    if root.exists():
        active_files = sorted(root.rglob("*.jsonl") if recursive else root.glob("*.jsonl"))
        archive_files = sorted(root.rglob("*.jsonl.gz") if recursive else root.glob("*.jsonl.archive/*.jsonl.gz"))
        files = [*archive_files, *active_files]
    else:
        files = []
    selected: dict[str, LoadedEvent] = {}
    lines_read = malformed = invalid = duplicates = sequence = 0

    for path in files:
        with lock_for(str(path)):
            opener = gzip.open if path.name.endswith(".jsonl.gz") else open
            with opener(path, mode="rt", encoding="utf-8") as handle:
                for line_number, line in enumerate(handle, start=1):
                    lines_read += 1
                    if not line.strip():
                        continue
                    try:
                        payload = json.loads(line)
                    except (json.JSONDecodeError, UnicodeError) as exc:
                        malformed += 1
                        LOGGER.warning("skip malformed JSONL: %s:%d (%s)", path, line_number, type(exc).__name__)
                        continue
                    if not isinstance(payload, dict):
                        invalid += 1
                        LOGGER.warning("skip non-object JSONL: %s:%d", path, line_number)
                        continue
                    try:
                        event = TokenEvent.from_dict(payload)
                    except (KeyError, TypeError, ValueError, AttributeError) as exc:
                        invalid += 1
                        LOGGER.warning("skip schema-invalid event: %s:%d (%s)", path, line_number, type(exc).__name__)
                        continue

                    sequence += 1
                    loaded = LoadedEvent(event, str(path.resolve()), line_number, sequence)
                    previous = selected.get(event.event_id)
                    if previous is not None:
                        duplicates += 1
                        # The immutable ledger's canonical identity rule is first event_id
                        # wins. A later row with the same identity is evidence of a collision,
                        # never a silent replacement of the source-of-truth event.
                        continue
                    selected[event.event_id] = loaded

    loaded_events = sorted(selected.values(), key=lambda item: item.sequence)
    projected = {event.event_id: event for event in effective_events([item.event for item in loaded_events])}
    loaded_events = [
        LoadedEvent(projected[item.event.event_id], item.source_file, item.source_line, item.sequence)
        for item in loaded_events
    ]
    report = LoadReport(
        files_read=len(files),
        lines_read=lines_read,
        valid_events=len(loaded_events),
        malformed_lines=malformed,
        schema_invalid_lines=invalid,
        duplicate_event_ids=duplicates,
    )
    return loaded_events, report


def load_prices(path: str | os.PathLike[str] | None) -> pd.DataFrame:
    """Load effective-dated prices. Missing prices stay unknown; they never become zero."""
    if path is None or not Path(path).exists():
        LOGGER.warning("price table not found; cost cells will remain unknown")
        return pd.DataFrame(columns=PRICE_COLUMNS)
    prices = pd.read_csv(path, dtype={"provider": "string", "model": "string", "token_type": "string", "currency": "string"})
    missing = [column for column in PRICE_COLUMNS if column not in prices.columns]
    if missing:
        raise ValueError(f"price table missing columns: {', '.join(missing)}")
    prices = prices.loc[:, PRICE_COLUMNS].copy()
    for column in ("provider", "model", "token_type", "currency"):
        prices[column] = prices[column].fillna("").str.strip()
    blank_dimensions = [
        column
        for column in ("provider", "model", "token_type", "currency")
        if (prices[column] == "").any()
    ]
    if blank_dimensions:
        raise ValueError(f"price table contains blank required fields: {', '.join(blank_dimensions)}")
    valid_token_types = {token_type.value for token_type in TokenType} | {"*"}
    invalid_token_types = sorted(set(prices.loc[~prices["token_type"].isin(valid_token_types), "token_type"]))
    if invalid_token_types:
        raise ValueError(f"price table contains unknown token_type values: {', '.join(invalid_token_types)}")
    prices["currency"] = prices["currency"].str.upper()
    invalid_currencies = sorted(
        value
        for value in set(prices["currency"])
        if len(value) != 3 or not value.isalpha()
    )
    if invalid_currencies:
        raise ValueError(f"currency must use three-letter codes: {', '.join(invalid_currencies)}")
    prices["price_per_million_tokens"] = pd.to_numeric(prices["price_per_million_tokens"], errors="coerce")
    if prices["price_per_million_tokens"].isna().any() or (prices["price_per_million_tokens"] < 0).any():
        raise ValueError("price_per_million_tokens must contain non-negative numbers")

    for column in ("effective_from", "effective_to"):
        parsed_dates: list[object] = []
        for row_number, value in enumerate(prices[column], start=2):
            text = "" if pd.isna(value) else str(value).strip()
            if not text:
                parsed_dates.append(pd.NaT)
                continue
            try:
                parsed_dates.append(pd.Timestamp(dt.date.fromisoformat(text)))
            except ValueError as exc:
                raise ValueError(
                    f"{column} row {row_number} must be YYYY-MM-DD or blank"
                ) from exc
        prices[column] = parsed_dates

    inverted = prices["effective_from"].notna() & prices["effective_to"].notna() & (
        prices["effective_to"] < prices["effective_from"]
    )
    if inverted.any():
        rows = ", ".join(str(index + 2) for index in prices.index[inverted])
        raise ValueError(f"effective_to precedes effective_from on rows: {rows}")

    # Two simultaneously active rows with the same selector would make cost depend on CSV
    # order. Reject the ambiguity instead of silently choosing one.
    for selector, group in prices.groupby(["provider", "model", "token_type"], sort=False):
        records = list(group.itertuples())
        for left_index, left in enumerate(records):
            left_start = pd.Timestamp.min if pd.isna(left.effective_from) else left.effective_from
            left_end = pd.Timestamp.max if pd.isna(left.effective_to) else left.effective_to
            for right in records[left_index + 1 :]:
                right_start = pd.Timestamp.min if pd.isna(right.effective_from) else right.effective_from
                right_end = pd.Timestamp.max if pd.isna(right.effective_to) else right.effective_to
                if left_start <= right_end and right_start <= left_end:
                    joined = "/".join(selector)
                    raise ValueError(f"overlapping effective price ranges for selector: {joined}")
    return prices


def _find_price(
    prices: pd.DataFrame,
    *,
    provider: str,
    model: str,
    token_type: str,
    event_date: dt.date | None,
) -> tuple[float, str] | None:
    if prices.empty:
        return None
    candidates = prices[
        prices["provider"].isin([provider, "*"])
        & prices["model"].isin([model, "*"])
        & prices["token_type"].isin([token_type, "*"])
    ].copy()
    if event_date is not None:
        event_timestamp = pd.Timestamp(event_date)
        candidates = candidates[
            (candidates["effective_from"].isna() | (candidates["effective_from"] <= event_timestamp))
            & (candidates["effective_to"].isna() | (candidates["effective_to"] >= event_timestamp))
        ]
    else:
        # A missing event timestamp cannot be matched safely to a dated tariff.
        candidates = candidates[candidates["effective_from"].isna() & candidates["effective_to"].isna()]
    if candidates.empty:
        return None
    candidates["_specificity"] = (
        (candidates["provider"] != "*").astype(int)
        + (candidates["model"] != "*").astype(int)
        + (candidates["token_type"] != "*").astype(int)
    )
    most_specific = candidates[candidates["_specificity"] == candidates["_specificity"].max()]
    if len(most_specific) != 1:
        selectors = sorted(
            f"{row.provider}/{row.model}/{row.token_type}"
            for row in most_specific.itertuples()
        )
        raise ValueError(f"ambiguous price match at equal specificity: {', '.join(selectors)}")
    selected = most_specific.iloc[0]
    return float(selected["price_per_million_tokens"]), str(selected["currency"])


def _billing_allocations(event: TokenEvent) -> tuple[list[int | None], list[str | None]]:
    """Allocate parent totals net of subtotals so cache/reasoning pricing cannot double count."""
    if event.superseded or not event.is_authoritative:
        return [0 for _ in event.quantities], ["excluded_event" for _ in event.quantities]

    allocations = [
        quantity.quantity if quantity.trust == Trust.VERIFIED else None
        for quantity in event.quantities
    ]
    issues: list[str | None] = [
        None if quantity.trust == Trust.VERIFIED else "unverified_additivity"
        for quantity in event.quantities
    ]
    parent_by_type: dict[str, int] = {}
    for index, quantity in enumerate(event.quantities):
        if quantity.subtotal_of is None and quantity.token_type.value not in parent_by_type:
            parent_by_type[quantity.token_type.value] = index

    children: dict[int, list[int]] = {}
    for index, quantity in enumerate(event.quantities):
        if not quantity.subtotal_of:
            continue
        parent_index = parent_by_type.get(quantity.subtotal_of)
        if parent_index is None:
            issues[index] = "orphan_subtotal"
            continue
        children.setdefault(parent_index, []).append(index)

    for parent_index, child_indices in children.items():
        family_indices = [parent_index, *child_indices]
        family = [event.quantities[index] for index in family_indices]
        if any(quantity.trust != Trust.VERIFIED for quantity in family):
            for index in family_indices:
                allocations[index] = None
                issues[index] = "unverified_subtotal_allocation"
            continue

        # Cache, reasoning, and modality details are separate decomposition axes. Two
        # axes can overlap (for example cached image input), so subtracting both from the
        # parent would fabricate a billable remainder. Refuse that allocation.
        dimensions = {
            "cache"
            if quantity.token_type == TokenType.CACHED_INPUT
            else "reasoning"
            if quantity.token_type in {TokenType.REASONING, TokenType.THINKING}
            else "modality"
            if quantity.token_type
            in {TokenType.AUDIO_INPUT, TokenType.AUDIO_OUTPUT, TokenType.IMAGE_INPUT, TokenType.VIDEO_INPUT}
            else quantity.token_type.value
            for quantity in family[1:]
        }
        if len(dimensions) > 1:
            for index in family_indices:
                allocations[index] = None
                issues[index] = "ambiguous_subtotal_overlap"
            continue

        parent_quantity = event.quantities[parent_index].quantity
        known_children = [event.quantities[index].quantity for index in child_indices]
        if parent_quantity is None or any(value is None for value in known_children):
            issues[parent_index] = "incomplete_subtotal_allocation"
            continue
        child_total = sum(int(value) for value in known_children if value is not None)
        if child_total > parent_quantity:
            return [None for _ in event.quantities], ["subtotal_exceeds_parent" for _ in event.quantities]
        allocations[parent_index] = parent_quantity - child_total
    return allocations, issues


def _latest_request_events(events: list[LoadedEvent]) -> dict[str, TokenEvent]:
    selected: dict[str, LoadedEvent] = {}
    for item in events:
        event = item.event
        if event.superseded or not event.is_authoritative:
            continue
        current = selected.get(event.request_correlation_id)
        if current is None or _newer(item, current):
            selected[event.request_correlation_id] = item
    return {correlation_id: item.event for correlation_id, item in selected.items()}


def build_data_frame(events: list[LoadedEvent], prices: pd.DataFrame) -> pd.DataFrame:
    """Flatten to quantity grain while exposing event/request values exactly once."""
    rows: list[dict[str, Any]] = []
    first_row_by_event: dict[str, int] = {}

    for item in events:
        event = item.event
        observation = event.observation
        deployment = observation.get("deployment")
        if not deployment:
            deployment = next(
                (
                    quantity.metadata.get("azure_deployment")
                    for quantity in event.quantities
                    if quantity.metadata.get("azure_deployment")
                ),
                None,
            )
        cloud_provider = observation.get("cloud_provider") or {
            "azure_openai": "azure",
            "bedrock": "aws",
            "vertex_ai": "gcp",
        }.get(event.provider or "", event.provider or "unknown")
        parsed_timestamp = _parse_timestamp(event.timestamp)
        excel_timestamp = parsed_timestamp.replace(tzinfo=None) if parsed_timestamp else None
        event_date = parsed_timestamp.date() if parsed_timestamp else None
        quantities = event.quantities or [None]
        allocations, allocation_issues = _billing_allocations(event) if event.quantities else ([None], ["no_quantity"])
        first_row_by_event[event.event_id] = len(rows)

        for quantity_index, quantity in enumerate(quantities):
            first_quantity = quantity_index == 0
            token_type = quantity.token_type.value if quantity else ""
            raw_tokens = quantity.quantity if quantity else None
            billing_tokens = allocations[quantity_index]
            allocation_issue = allocation_issues[quantity_index]
            price = _find_price(
                prices,
                provider=event.provider or "unknown",
                model=event.model or "unknown",
                token_type=token_type,
                event_date=event_date,
            ) if quantity else None

            if event.superseded or not event.is_authoritative:
                unit_price = None
                currency = None
                derived_cost = 0.0
                cost_quality = "excluded_event"
            elif billing_tokens is None:
                unit_price = price[0] if price else None
                currency = price[1] if price else None
                derived_cost = None
                cost_quality = allocation_issue or "unknown_quantity"
            elif billing_tokens == 0:
                unit_price = price[0] if price else None
                currency = price[1] if price else None
                derived_cost = 0.0
                cost_quality = allocation_issue or "zero_component"
            elif price is None:
                unit_price = None
                currency = None
                derived_cost = None
                cost_quality = "missing_price"
            else:
                unit_price, currency = price
                # Reporting rule: cost = allocated billing tokens x unit price per million.
                derived_cost = billing_tokens * unit_price / 1_000_000
                uncertain = quantity.precision_level.value != "exact" or quantity.trust.value != "verified"
                cost_quality = allocation_issue or ("estimated" if uncertain else "exact")

            active_event = event.is_authoritative and not event.superseded
            active_quantity_in_total = (
                quantity.quantity_in_total if quantity and active_event else 0
            )
            billable_for_coverage = billing_tokens if active_event and billing_tokens is not None and billing_tokens > 0 else 0
            priced_billing_tokens = billable_for_coverage if derived_cost is not None else 0
            cache_read_tokens = (
                raw_tokens
                if active_event and token_type == "cached_input" and raw_tokens is not None
                else 0
            )

            rows.append(
                {
                    "source_file": item.source_file,
                    "source_line": item.source_line,
                    "source_kind": observation.get("source") or "unspecified",
                    "source_event_id": observation.get("source_event_id"),
                    "event_id": event.event_id,
                    "request_correlation_id": event.request_correlation_id,
                    "trace_id": event.trace_id,
                    "span_id": event.span_id,
                    "parent_span_id": event.parent_span_id,
                    "business_id": event.business_id,
                    "use_case": observation.get("use_case") or event.workflow or "unknown",
                    "workflow": event.workflow or "unknown",
                    "environment": event.environment or "unknown",
                    "timestamp_utc": excel_timestamp,
                    "date": event_date,
                    "provider": event.provider or "unknown",
                    "model": event.model or "unknown",
                    "deployment": deployment or "unknown",
                    "cloud_provider": cloud_provider,
                    "region": observation.get("region") or "unknown",
                    "service_name": observation.get("service_name") or "unknown",
                    "tenant_id": observation.get("tenant_id") or "unknown",
                    "provider_request_id": observation.get("provider_request_id"),
                    "provider_response_id": observation.get("provider_response_id"),
                    "retry_count": observation.get("retry_count") or 0,
                    "api_surface": event.api_surface or "unknown",
                    "event_authoritative": event.is_authoritative,
                    "event_status": observation.get("status") or "unknown",
                    "http_status": observation.get("http_status"),
                    "event_superseded": event.superseded,
                    "superseded_by": event.superseded_by,
                    "data_quality_flags": ",".join(event.data_quality_flags),
                    "token_type": token_type,
                    "raw_tokens": raw_tokens,
                    "precision_level": quantity.precision_level.value if quantity else "unknown",
                    "usage_source": quantity.usage_source.value if quantity else "none",
                    "overlap": quantity.overlap.value if quantity else "",
                    "trust": quantity.trust.value if quantity else "",
                    "subtotal_of": quantity.subtotal_of if quantity else None,
                    "quantity_in_total": active_quantity_in_total,
                    "exact_tokens_active": (
                        active_quantity_in_total
                        if quantity and quantity.precision_level.value == "exact"
                        else 0
                    ),
                    "estimated_tokens_active": (
                        active_quantity_in_total
                        if quantity and quantity.precision_level.value == "estimate"
                        else 0
                    ),
                    "unverified_tokens_active": (
                        raw_tokens
                        if quantity
                        and active_event
                        and quantity.trust == Trust.UNVERIFIED
                        and quantity.overlap.value == "independent"
                        and raw_tokens is not None
                        else 0
                    ),
                    "event_contributing_tokens_once": event.event_contributing_tokens if first_quantity else 0,
                    "event_count_once": 1 if first_quantity else 0,
                    "event_authoritative_once": 1 if first_quantity and event.is_authoritative else 0,
                    "quality_flagged_event_once": 1 if first_quantity and event.data_quality_flags else 0,
                    "active_quality_flagged_event_once": (
                        1 if first_quantity and active_event and event.data_quality_flags else 0
                    ),
                    "superseded_event_once": 1 if first_quantity and event.superseded else 0,
                    "mismatch_event_once": (
                        1
                        if first_quantity
                        and active_event
                        and event.event_total_mismatch not in (None, 0)
                        else 0
                    ),
                    "provider_total_tokens_once": event.provider_total_tokens if first_quantity else None,
                    "provider_total_observation_once": (
                        1 if first_quantity and active_event and event.provider_total_tokens is not None else 0
                    ),
                    "event_total_mismatch_once": event.event_total_mismatch if first_quantity else None,
                    "under_attributed_tokens_once": event.under_attributed_tokens if first_quantity else 0,
                    "over_attributed_tokens_once": event.over_attributed_tokens if first_quantity else 0,
                    "schema_drift_event_once": (
                        1
                        if first_quantity
                        and active_event
                        and DataQualityFlag.PROVIDER_SCHEMA_DRIFT.value in event.data_quality_flags
                        else 0
                    ),
                    "correlation_risk_event_once": (
                        1
                        if first_quantity
                        and {
                            DataQualityFlag.CORRELATION_ID_COLLISION.value,
                            DataQualityFlag.DUPLICATE_FINAL_UNVERIFIED.value,
                        }
                        & set(event.data_quality_flags)
                        else 0
                    ),
                    "usage_loss_event_once": (
                        1
                        if first_quantity and active_event and USAGE_LOSS_FLAGS & set(event.data_quality_flags)
                        else 0
                    ),
                    "billing_tokens": billing_tokens,
                    "billable_tokens_for_coverage": billable_for_coverage,
                    "priced_billing_tokens": priced_billing_tokens,
                    "cache_read_tokens_active": cache_read_tokens,
                    "unknown_quantity_active": (
                        1
                        if active_event and (quantity is None or quantity.precision_level.value == "unknown")
                        else 0
                    ),
                    "unit_price_per_million": unit_price,
                    "currency": currency,
                    "derived_cost": derived_cost,
                    "cost_quality": cost_quality,
                    "request_count_once": 0,
                    "request_latency_applicable_once": 0,
                    "request_latency_ms": None,
                    "request_latency_observation_once": 0,
                    "request_ttft_applicable_once": 0,
                    "request_ttft_ms": None,
                    "request_ttft_observation_once": 0,
                    "quantity_metadata_json": json.dumps(quantity.metadata, ensure_ascii=True, sort_keys=True) if quantity else "{}",
                    # Observation is event-grain. Repeating it for every quantity can multiply
                    # workbook size without adding information; the first quantity row remains
                    # the auditable event record.
                    "observation_json": (
                        json.dumps(observation.to_dict(), ensure_ascii=True, sort_keys=True)
                        if first_quantity
                        else None
                    ),
                }
            )

    latest_requests = _latest_request_events(events)
    for event in latest_requests.values():
        row_index = first_row_by_event[event.event_id]
        flags = set(event.data_quality_flags)
        local_log_import = bool(
            flags
            & {
                DataQualityFlag.CLAUDE_CODE_LOCAL_USAGE.value,
                DataQualityFlag.CODEX_LOCAL_TOKEN_COUNT.value,
            }
        )
        stream_request = any(
            quantity.usage_source.value.startswith("provider_stream_")
            or quantity.usage_source.value == "partial_stream_tokenizer"
            for quantity in event.quantities
        )
        rows[row_index]["request_count_once"] = 1
        rows[row_index]["request_latency_applicable_once"] = int(not local_log_import)
        # Latency rule: one duration per request_correlation_id, never one per quantity/event row.
        rows[row_index]["request_latency_ms"] = event.observation.get("duration_ms")
        rows[row_index]["request_latency_observation_once"] = int(event.observation.get("duration_ms") is not None)
        rows[row_index]["request_ttft_applicable_once"] = int(stream_request and not local_log_import)
        rows[row_index]["request_ttft_ms"] = event.observation.get("time_to_first_token_ms")
        rows[row_index]["request_ttft_observation_once"] = int(
            event.observation.get("time_to_first_token_ms") is not None
        )

    return pd.DataFrame(rows, columns=DATA_COLUMNS)


def _ratio_or_none(numerator: int | float, denominator: int | float) -> float | None:
    return float(numerator / denominator) if denominator else None


def _quality_and_provenance_frames(data: pd.DataFrame) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame]:
    event_rows = data[data["event_count_once"] == 1]
    active_events = event_rows[event_rows["event_authoritative"] & ~event_rows["event_superseded"]]
    request_rows = data[data["request_count_once"] == 1]

    exact_tokens = int(data["exact_tokens_active"].sum())
    estimated_tokens = int(data["estimated_tokens_active"].sum())
    unverified_tokens = int(data["unverified_tokens_active"].sum())
    known_token_magnitude = exact_tokens + estimated_tokens + unverified_tokens
    request_count = int(request_rows["request_count_once"].sum())
    latency_applicable_requests = int(request_rows["request_latency_applicable_once"].sum())
    latency_observations = int(request_rows["request_latency_observation_once"].sum())
    ttft_applicable_requests = int(request_rows["request_ttft_applicable_once"].sum())
    ttft_observations = int(request_rows["request_ttft_observation_once"].sum())
    active_event_count = len(active_events)
    provider_total_observations = int(active_events["provider_total_observation_once"].sum())
    billable_tokens = int(data["billable_tokens_for_coverage"].sum())
    priced_tokens = int(data["priced_billing_tokens"].sum())

    quality = {
        "data_row_count": len(data),
        "volume_status": "warning" if len(data) >= DASHBOARD_ROW_WARNING else "ok",
        "event_count": len(event_rows),
        "active_event_count": active_event_count,
        "excluded_event_count": len(event_rows) - active_event_count,
        "request_count": request_count,
        "source_file_count": int(data["source_file"].nunique()),
        "source_kind_count": int(data["source_kind"].nunique()),
        "exact_tokens": exact_tokens,
        "estimated_tokens": estimated_tokens,
        "unverified_tokens": unverified_tokens,
        "known_exact_token_share": _ratio_or_none(exact_tokens, known_token_magnitude),
        "unknown_quantity_count": int(data["unknown_quantity_active"].sum()),
        "active_flagged_event_count": int(active_events["active_quality_flagged_event_once"].sum()),
        "mismatch_event_count": int(active_events["mismatch_event_once"].sum()),
        "schema_drift_event_count": int(active_events["schema_drift_event_once"].sum()),
        # Correlation collisions frequently become superseded during reconciliation. They
        # remain integrity failures and must not disappear from the audit summary.
        "correlation_risk_event_count": int(event_rows["correlation_risk_event_once"].sum()),
        "usage_loss_event_count": int(active_events["usage_loss_event_once"].sum()),
        "provider_total_coverage": _ratio_or_none(provider_total_observations, active_event_count),
        "latency_coverage": _ratio_or_none(latency_observations, request_count),
        "instrumented_latency_coverage": _ratio_or_none(latency_observations, latency_applicable_requests),
        "latency_applicability": _ratio_or_none(latency_applicable_requests, request_count),
        "ttft_coverage": _ratio_or_none(ttft_observations, ttft_applicable_requests),
        "ttft_applicability": _ratio_or_none(ttft_applicable_requests, request_count),
        "pricing_coverage": _ratio_or_none(priced_tokens, billable_tokens),
    }
    blocking_quality = (
        quality["usage_loss_event_count"]
        + quality["schema_drift_event_count"]
        + quality["correlation_risk_event_count"]
        + int(data["over_attributed_tokens_once"].sum() > 0)
    )
    uncertain_quality = (
        estimated_tokens
        + unverified_tokens
        + quality["unknown_quantity_count"]
        + quality["mismatch_event_count"]
        + int(data["under_attributed_tokens_once"].sum() > 0)
    )
    quality["quality_status"] = "blocked" if blocking_quality else ("warning" if uncertain_quality else "clean")
    required_coverage = [quality["pricing_coverage"], quality["instrumented_latency_coverage"]]
    observed_coverage = [value for value in required_coverage if value is not None]
    if not observed_coverage or all(value == 0 for value in observed_coverage):
        quality["coverage_status"] = "missing"
    elif len(observed_coverage) == len(required_coverage) and all(value == 1.0 for value in observed_coverage):
        quality["coverage_status"] = "complete"
    else:
        quality["coverage_status"] = "partial"

    aggregate_columns = {
        "Events": ("event_count_once", "sum"),
        "Authoritative events": ("event_authoritative_once", "sum"),
        "Superseded events": ("superseded_event_once", "sum"),
        "Requests": ("request_count_once", "sum"),
        "Latency-applicable requests": ("request_latency_applicable_once", "sum"),
        "Contributing tokens": ("event_contributing_tokens_once", "sum"),
        "Exact tokens": ("exact_tokens_active", "sum"),
        "Estimated tokens": ("estimated_tokens_active", "sum"),
        "Unverified tokens": ("unverified_tokens_active", "sum"),
        "Unknown quantities": ("unknown_quantity_active", "sum"),
        "Flagged events": ("active_quality_flagged_event_once", "sum"),
        "Usage-loss events": ("usage_loss_event_once", "sum"),
        "Provider-total observations": ("provider_total_observation_once", "sum"),
        "Latency observations": ("request_latency_observation_once", "sum"),
        "Billable tokens": ("billable_tokens_for_coverage", "sum"),
        "Priced tokens": ("priced_billing_tokens", "sum"),
    }

    def summarize_by(keys: list[str]) -> pd.DataFrame:
        grouped = data.groupby(keys, as_index=False, dropna=False).agg(**aggregate_columns)
        grouped["Active events"] = grouped["Authoritative events"] - grouped["Superseded events"]
        grouped["Provider-total coverage"] = grouped.apply(
            lambda row: _ratio_or_none(row["Provider-total observations"], row["Active events"]), axis=1
        )
        grouped["Latency coverage"] = grouped.apply(
            lambda row: _ratio_or_none(row["Latency observations"], row["Requests"]), axis=1
        )
        grouped["Instrumented latency coverage"] = grouped.apply(
            lambda row: _ratio_or_none(row["Latency observations"], row["Latency-applicable requests"]), axis=1
        )
        grouped["Latency applicability"] = grouped.apply(
            lambda row: _ratio_or_none(row["Latency-applicable requests"], row["Requests"]), axis=1
        )
        grouped["Pricing coverage"] = grouped.apply(
            lambda row: _ratio_or_none(row["Priced tokens"], row["Billable tokens"]), axis=1
        )
        return grouped.sort_values("Contributing tokens", ascending=False)

    provider_summary = summarize_by(["cloud_provider", "provider", "api_surface"])
    source_summary = summarize_by(["source_kind", "source_file"])
    return quality, provider_summary, source_summary


def build_summary_frames(data: pd.DataFrame) -> dict[str, Any]:
    if data.empty:
        empty = pd.DataFrame()
        return {
            "currency": None,
            "total_cost": None,
            "pricing_coverage": None,
            "cost_by_day": empty,
            "cost_by_model": empty,
            "tokens_by_day": empty,
            "latency_by_day_model": empty,
            "use_cases": empty,
            "quality": {
                "data_row_count": 0,
                "volume_status": "ok",
                "quality_status": "clean",
            },
            "provider_summary": empty,
            "source_summary": empty,
        }

    # A zero-token component can have a mathematically known zero cost even when every
    # token-bearing component lacks a price. It must not turn an entirely unpriced workload
    # into a misleading known total of 0. Only positively billable, priced rows participate
    # in known-cost summaries; pricing_coverage below quantifies the omitted magnitude.
    cost_rows = data[
        data["derived_cost"].notna()
        & (data["billing_tokens"].fillna(0) > 0)
        & (data["cost_quality"] != "excluded_event")
    ]
    currencies = sorted(value for value in cost_rows["currency"].dropna().unique() if value)
    if len(currencies) > 1:
        raise ValueError("dashboard cannot aggregate multiple currencies; filter the price table to one currency")
    currency = currencies[0] if currencies else None
    total_cost = float(cost_rows["derived_cost"].sum()) if not cost_rows.empty else None

    billable = data[(data["billing_tokens"].fillna(0) > 0) & ~data["event_superseded"] & data["event_authoritative"]]
    priced_tokens = billable.loc[billable["derived_cost"].notna(), "billing_tokens"].sum()
    all_tokens = billable["billing_tokens"].sum()
    pricing_coverage = float(priced_tokens / all_tokens) if all_tokens else None

    dated_costs = cost_rows[cost_rows["date"].notna()]
    cost_by_day = (
        dated_costs.groupby("date", as_index=False)["derived_cost"].sum().sort_values("date")
        if not dated_costs.empty
        else pd.DataFrame(columns=["date", "derived_cost"])
    )
    cost_by_model = (
        cost_rows.groupby("model", as_index=False)["derived_cost"].sum().sort_values("derived_cost", ascending=False)
        if not cost_rows.empty
        else pd.DataFrame(columns=["model", "derived_cost"])
    )

    token_rows = data[data["date"].notna() & (data["quantity_in_total"] > 0)]
    tokens_by_day = (
        token_rows.pivot_table(index="date", columns="token_type", values="quantity_in_total", aggfunc="sum", fill_value=0)
        .reset_index()
        .sort_values("date")
        if not token_rows.empty
        else pd.DataFrame(columns=["date"])
    )

    requests = data[(data["request_count_once"] == 1) & data["date"].notna() & data["request_latency_ms"].notna()].copy()
    latency = (
        requests.groupby(["date", "model"])["request_latency_ms"]
        .agg(average_latency_ms="mean", p95_latency_ms=lambda values: values.quantile(0.95))
        .reset_index()
        if not requests.empty
        else pd.DataFrame(columns=["date", "model", "average_latency_ms", "p95_latency_ms"])
    )

    use_case_usage = data.groupby("use_case", as_index=False).agg(
        contributing_tokens=("event_contributing_tokens_once", "sum"),
        requests=("request_count_once", "sum"),
    )
    use_case_cost = cost_rows.groupby("use_case", as_index=False)["derived_cost"].sum()
    use_cases = use_case_usage.merge(use_case_cost, on="use_case", how="left").sort_values(
        ["derived_cost", "contributing_tokens"],
        ascending=[False, False],
        na_position="last",
    )
    quality, provider_summary, source_summary = _quality_and_provenance_frames(data)
    return {
        "currency": currency,
        "total_cost": total_cost,
        "pricing_coverage": pricing_coverage,
        "cost_by_day": cost_by_day,
        "cost_by_model": cost_by_model,
        "tokens_by_day": tokens_by_day,
        "latency_by_day_model": latency,
        "use_cases": use_cases,
        "quality": quality,
        "provider_summary": provider_summary,
        "source_summary": source_summary,
    }


def _excel_value(value: Any) -> Any:
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _write_frame(ws, frame: pd.DataFrame, *, start_row: int, start_col: int, table_name: str) -> tuple[int, int]:
    headers = list(frame.columns)
    for column_offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + column_offset, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_offset, values in enumerate(frame.itertuples(index=False, name=None), start=1):
        for column_offset, value in enumerate(values):
            ws.cell(start_row + row_offset, start_col + column_offset, _excel_value(value))
    end_row = start_row + max(len(frame), 1)
    end_col = start_col + max(len(headers), 1) - 1
    if headers and not frame.empty:
        reference = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    elif headers:
        ws.cell(start_row + 1, start_col, "No data")
    return end_row, end_col


def _style_dashboard(ws, title: str, subtitle: str, *, end_col: int = 8) -> None:
    ws.sheet_view.showGridLines = False
    end_letter = get_column_letter(end_col)
    ws.merge_cells(f"A1:{end_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{end_letter}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=MUTED)
    ws.row_dimensions[2].height = 24


def _write_kpi(ws, label_cell: str, value_cell: str, label: str, value: Any, number_format: str) -> None:
    ws[label_cell] = label
    ws[label_cell].font = Font(size=9, bold=True, color=MUTED)
    ws[value_cell] = value
    ws[value_cell].font = Font(size=18, bold=True, color=NAVY)
    ws[value_cell].number_format = number_format
    for cell in (ws[label_cell], ws[value_cell]):
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        cell.border = Border(bottom=Side(style="thin", color=GRID))


def _criteria_arguments(
    *,
    overrides: dict[str, str] | None = None,
    extra: tuple[tuple[str, str], ...] = (),
    include_dates: bool = True,
) -> str:
    overrides = overrides or {}
    arguments: list[str] = []
    for column, reference in DASHBOARD_FILTER_REFS.items():
        criterion = overrides.get(column, f'IF({reference}="All","*",{reference})')
        arguments.extend((f"DataTable[{column}]", criterion))
    if "date" in overrides:
        arguments.extend(("DataTable[date]", overrides["date"]))
    elif include_dates:
        arguments.extend(
            (
                "DataTable[date]",
                f'">="&{DASHBOARD_DATE_FROM}',
                "DataTable[date]",
                f'"<="&{DASHBOARD_DATE_TO}',
            )
        )
    for column, criterion in extra:
        arguments.extend((f"DataTable[{column}]", criterion))
    return ",".join(arguments)


def _sumifs_formula(
    column: str,
    *,
    overrides: dict[str, str] | None = None,
    extra: tuple[tuple[str, str], ...] = (),
    include_dates: bool = True,
) -> str:
    criteria = _criteria_arguments(overrides=overrides, extra=extra, include_dates=include_dates)
    return f"IFERROR(SUMIFS(DataTable[{column}],{criteria}),0)"


def _ratio_sumifs_formula(
    numerator_column: str,
    denominator_column: str,
    *,
    overrides: dict[str, str] | None = None,
    include_dates: bool = True,
) -> str:
    numerator = _sumifs_formula(numerator_column, overrides=overrides, include_dates=include_dates)
    denominator = _sumifs_formula(denominator_column, overrides=overrides, include_dates=include_dates)
    return f'IFERROR(({numerator})/({denominator}),"")'


def _filter_mask(
    *,
    overrides: dict[str, str] | None = None,
    extra: tuple[tuple[str, str], ...] = (),
    include_dates: bool = True,
) -> str:
    overrides = overrides or {}
    conditions: list[str] = []
    for column, reference in DASHBOARD_FILTER_REFS.items():
        if column in overrides:
            conditions.append(f"(DataTable[{column}]={overrides[column]})")
        else:
            conditions.append(f'(((DataTable[{column}]={reference})+({reference}="All"))>0)')
    if "date" in overrides:
        conditions.append(f"(DataTable[date]={overrides['date']})")
    elif include_dates:
        conditions.extend(
            (
                f"(DataTable[date]>={DASHBOARD_DATE_FROM})",
                f"(DataTable[date]<={DASHBOARD_DATE_TO})",
            )
        )
    conditions.extend(f"(DataTable[{column}]={criterion})" for column, criterion in extra)
    return "*".join(conditions)


def _p95_formula(
    column: str,
    *,
    overrides: dict[str, str] | None = None,
    include_dates: bool = True,
) -> str:
    mask = _filter_mask(
        overrides=overrides,
        extra=(("request_count_once", "1"),),
        include_dates=include_dates,
    )
    return (
        f'IFERROR(PERCENTILE.INC(_xlfn._xlws.FILTER(DataTable[{column}],'
        f'{mask}*ISNUMBER(DataTable[{column}])),0.95),"")'
    )


def _write_dashboard_lists(workbook: Workbook, ws_lists, data: pd.DataFrame) -> tuple[dt.date, dt.date]:
    ws_lists.sheet_state = "veryHidden"
    for column_index, (label, data_column, _, range_name) in enumerate(DASHBOARD_FILTERS, start=1):
        values = ["All"]
        if not data.empty:
            values.extend(sorted({str(value) for value in data[data_column].dropna() if str(value)}))
        ws_lists.cell(1, column_index, label)
        for row_index, value in enumerate(values, start=2):
            ws_lists.cell(row_index, column_index, value)
        column_letter = get_column_letter(column_index)
        workbook.defined_names.add(
            DefinedName(
                range_name,
                attr_text=f"'{ws_lists.title}'!${column_letter}$2:${column_letter}${len(values) + 1}",
            )
        )

    dates = sorted(value for value in data["date"].dropna().unique()) if not data.empty else []
    if not dates:
        today = dt.datetime.now(dt.UTC).date()
        return today, today
    chart_dates = dates[-MAX_DASHBOARD_DATES:]
    return chart_dates[0], chart_dates[-1]


def _style_filter_cell(ws, cell: str) -> None:
    ws[cell].fill = PatternFill("solid", fgColor=SOFT_BLUE)
    ws[cell].font = Font(color=NAVY, bold=True)
    ws[cell].alignment = Alignment(horizontal="center")
    ws[cell].border = Border(bottom=Side(style="medium", color=AZURE_BLUE))


def _write_dashboard_filters(
    ws,
    *,
    minimum_date: dt.date,
    maximum_date: dt.date,
) -> None:
    for label, _, cell, range_name in DASHBOARD_FILTERS:
        label_cell = f"{cell[0]}4"
        ws[label_cell] = label
        ws[label_cell].font = Font(size=9, bold=True, color=MUTED)
        ws[cell] = "All"
        _style_filter_cell(ws, cell)
        validation = DataValidation(type="list", formula1=f"={range_name}", allow_blank=False)
        validation.error = "Choose a value from the tracker dimensions."
        validation.errorTitle = "Invalid filter"
        ws.add_data_validation(validation)
        validation.add(ws[cell])

    for label, cell, value in (("From", "L5", minimum_date), ("To", "N5", maximum_date)):
        ws[f"{cell[0]}4"] = label
        ws[f"{cell[0]}4"].font = Font(size=9, bold=True, color=MUTED)
        ws[cell] = value
        ws[cell].number_format = "yyyy-mm-dd"
        _style_filter_cell(ws, cell)
        validation = DataValidation(
            type="date",
            operator="between",
            formula1=f"DATE({minimum_date.year},{minimum_date.month},{minimum_date.day})",
            formula2=f"DATE({maximum_date.year},{maximum_date.month},{maximum_date.day})",
            allow_blank=False,
        )
        ws.add_data_validation(validation)
        validation.add(ws[cell])


def _write_dashboard_kpi(
    ws,
    *,
    start_col: int,
    start_row: int,
    label: str,
    formula: str,
    number_format: str,
    accent: str,
) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + 1)
    label_cell = ws.cell(start_row, start_col)
    value_cell = ws.cell(start_row + 1, start_col)
    label_cell.value = label
    label_cell.font = Font(size=9, bold=True, color=MUTED)
    label_cell.fill = PatternFill("solid", fgColor=LIGHT)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.value = f"={formula}"
    value_cell.font = Font(size=18, bold=True, color=NAVY)
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.number_format = number_format
    edge = Side(style="medium", color=accent)
    label_cell.border = Border(left=edge, top=Side(style="thin", color=GRID), right=Side(style="thin", color=GRID))
    value_cell.border = Border(left=edge, bottom=Side(style="thin", color=GRID), right=Side(style="thin", color=GRID))


def _write_dashboard_kpis(ws, *, has_data: bool) -> None:
    raw_cost = _sumifs_formula("derived_cost")
    tokens = _sumifs_formula("event_contributing_tokens_once")
    exact_tokens = _sumifs_formula("exact_tokens_active")
    estimated_tokens = _sumifs_formula("estimated_tokens_active")
    unverified_tokens = _sumifs_formula("unverified_tokens_active")
    requests = _sumifs_formula("request_count_once")
    average_latency = _ratio_sumifs_formula("request_latency_ms", "request_latency_observation_once")
    p95_latency = _p95_formula("request_latency_ms")
    priced = _sumifs_formula("priced_billing_tokens")
    billable = _sumifs_formula("billable_tokens_for_coverage")
    latency_observations = _sumifs_formula("request_latency_observation_once")
    unknown = _sumifs_formula("unknown_quantity_active")
    under_attributed = _sumifs_formula("under_attributed_tokens_once")
    over_attributed = _sumifs_formula("over_attributed_tokens_once")
    schema_drift = _sumifs_formula("schema_drift_event_once")
    correlation_risks = _sumifs_formula("correlation_risk_event_once")

    # An unknown token magnitude makes both cost and pricing coverage unknowable. Never
    # collapse that uncertainty into a visually reassuring zero.
    cost = f'IF(({unknown})>0,"N/A",IF(({billable})=0,0,IF(({priced})=0,"N/A",{raw_cost})))'
    pricing_coverage = f'IF(({unknown})>0,"N/A",IF(({billable})=0,"N/A",({priced})/({billable})))'
    latency_coverage = f'IF(({requests})=0,"N/A",({latency_observations})/({requests}))'

    primary = (
        ("Known cost", cost, "0.000000", AZURE_BLUE),
        ("Contributing tokens", tokens, "#,##0", AZURE_BLUE),
        ("Requests", requests, "#,##0", FOUNDRY_TEAL),
        ("Average latency", average_latency, '0.0 "ms"', FOUNDRY_TEAL),
        ("P95 latency", p95_latency, '0.0 "ms"', WARNING),
        ("Pricing coverage", pricing_coverage, "0.0%", SUCCESS),
        ("Latency coverage", latency_coverage, "0.0%", SUCCESS),
    )
    quality = (
        (
            "Known exact token share",
            f"IFERROR(({exact_tokens})/(({exact_tokens})+({estimated_tokens})+({unverified_tokens})),0)",
            "0.0%",
            SUCCESS,
        ),
        ("Estimated tokens", estimated_tokens, "#,##0", WARNING),
        ("Unknown quantities", unknown, "#,##0", WARNING),
        ("Under-attributed", under_attributed, "#,##0", WARNING),
        ("Over-attributed", over_attributed, "#,##0", CORAL),
        ("Schema drift events", schema_drift, "#,##0", CORAL),
        ("Correlation risks", correlation_risks, "#,##0", CORAL),
    )
    for index, (label, formula, number_format, accent) in enumerate(primary):
        _write_dashboard_kpi(
            ws,
            start_col=1 + index * 2,
            start_row=7,
            label=label,
            formula=formula if has_data else "0",
            number_format=number_format,
            accent=accent,
        )
    for index, (label, formula, number_format, accent) in enumerate(quality):
        _write_dashboard_kpi(
            ws,
            start_col=1 + index * 2,
            start_row=11,
            label=label,
            formula=formula if has_data else "0",
            number_format=number_format,
            accent=accent,
        )

    ws.merge_cells("A15:N15")
    ws["A15"] = (
        '=IF(E8=0,"NO REQUESTS IN FILTER",'
        'IF(OR(E12>0,G12>0,I12>0,K12>0,M12>0),"QUALITY REVIEW REQUIRED",'
        'IF(AND(ISNUMBER(K8),K8<1),"COST COVERAGE INCOMPLETE",'
        'IF(AND(ISNUMBER(M8),M8<1),"LATENCY COVERAGE INCOMPLETE","READY WITHIN FILTERED DATA"))))'
    )
    ws["A15"].font = Font(size=10, bold=True, color=NAVY)
    ws["A15"].fill = PatternFill("solid", fgColor=SOFT_BLUE)
    ws["A15"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A15"].border = Border(bottom=Side(style="thin", color=GRID))
    ws.row_dimensions[15].height = 22


def _write_dashboard_helper_tables(ws, data: pd.DataFrame) -> tuple[int, int]:
    dates = sorted(value for value in data["date"].dropna().unique())[-MAX_DASHBOARD_DATES:]
    model_order = (
        data.groupby("model", as_index=False)
        .agg(tokens=("event_contributing_tokens_once", "sum"), cost=("derived_cost", "sum"))
        .sort_values(["cost", "tokens"], ascending=False)["model"]
        .head(MAX_DASHBOARD_MODELS)
        .tolist()
    )
    start_row = 50
    ws.cell(start_row - 2, 1, "Calculated chart tables")
    ws.cell(start_row - 2, 1).font = Font(size=11, bold=True, color=NAVY)
    daily_headers = ("Date", "Date label", "Cost", "Tokens", "Requests", "Average latency", "P95 latency")
    for column, header in enumerate(daily_headers, start=1):
        ws.cell(start_row, column, header)
    for offset, date_value in enumerate(dates, start=1):
        row = start_row + offset
        date_ref = f"$A{row}"
        visibility = f"AND({date_ref}>={DASHBOARD_DATE_FROM},{date_ref}<={DASHBOARD_DATE_TO})"
        ws.cell(row, 1, date_value)
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 2, date_value.strftime("%Y-%m-%d"))
        ws.cell(
            row,
            3,
            f'=IF({visibility},{_sumifs_formula("derived_cost", overrides={"date": date_ref}, include_dates=False)},NA())',
        )
        ws.cell(
            row,
            4,
            "=IF("
            + visibility
            + ","
            + _sumifs_formula("event_contributing_tokens_once", overrides={"date": date_ref}, include_dates=False)
            + ",NA())",
        )
        ws.cell(
            row,
            5,
            f'=IF({visibility},{_sumifs_formula("request_count_once", overrides={"date": date_ref}, include_dates=False)},NA())',
        )
        ws.cell(
            row,
            6,
            "=IF("
            + visibility
            + ","
            + _ratio_sumifs_formula(
                "request_latency_ms",
                "request_latency_observation_once",
                overrides={"date": date_ref},
                include_dates=False,
            )
            + ",NA())",
        )
        ws.cell(
            row,
            7,
            f'=IF({visibility},{_p95_formula("request_latency_ms", overrides={"date": date_ref}, include_dates=False)},NA())',
        )

    model_start_col = 9
    model_headers = ("Model", "Cost", "Tokens", "Requests", "Average latency")
    for offset, header in enumerate(model_headers):
        ws.cell(start_row, model_start_col + offset, header)
    model_filter = DASHBOARD_FILTER_REFS["model"]
    for offset, model in enumerate(model_order, start=1):
        row = start_row + offset
        model_ref = f"$I{row}"
        visibility = f'OR({model_filter}="All",{model_filter}={model_ref})'
        ws.cell(row, model_start_col, model)
        metrics = (
            _sumifs_formula("derived_cost", overrides={"model": model_ref}),
            _sumifs_formula("event_contributing_tokens_once", overrides={"model": model_ref}),
            _sumifs_formula("request_count_once", overrides={"model": model_ref}),
            _ratio_sumifs_formula(
                "request_latency_ms",
                "request_latency_observation_once",
                overrides={"model": model_ref},
            ),
        )
        for metric_offset, formula in enumerate(metrics, start=1):
            ws.cell(row, model_start_col + metric_offset, f'=IF({visibility},{formula},NA())')

    for start_col, end_col, rows, name in (
        (1, len(daily_headers), len(dates), "DashboardDaily"),
        (model_start_col, model_start_col + len(model_headers) - 1, len(model_order), "DashboardModelSummary"),
    ):
        if not rows:
            continue
        for column in range(start_col, end_col + 1):
            cell = ws.cell(start_row, column)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(color=WHITE, bold=True)
        table = Table(
            displayName=name,
            ref=f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row + rows}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    return len(dates), len(model_order)


def _add_interactive_dashboard_charts(ws, daily_rows: int, model_rows: int) -> None:
    if daily_rows:
        cost = AreaChart()
        cost.title = "Cost by day"
        cost.y_axis.title = "Cost"
        cost.add_data(Reference(ws, min_col=3, min_row=50, max_row=50 + daily_rows), titles_from_data=True)
        cost.set_categories(Reference(ws, min_col=2, min_row=51, max_row=50 + daily_rows))
        cost.height = 7
        cost.width = 13
        cost.legend = None
        cost.style = 13
        cost.series[0].graphicalProperties.solidFill = AZURE_BLUE
        ws.add_chart(cost, "A16")

        tokens = BarChart()
        tokens.type = "col"
        tokens.title = "Tokens by day"
        tokens.y_axis.title = "Tokens"
        tokens.add_data(Reference(ws, min_col=4, min_row=50, max_row=50 + daily_rows), titles_from_data=True)
        tokens.set_categories(Reference(ws, min_col=2, min_row=51, max_row=50 + daily_rows))
        tokens.height = 7
        tokens.width = 13
        tokens.legend = None
        tokens.style = 10
        tokens.series[0].graphicalProperties.solidFill = FOUNDRY_TEAL
        ws.add_chart(tokens, "H16")

        latency = LineChart()
        latency.title = "Request latency by day"
        latency.y_axis.title = "Milliseconds"
        latency.add_data(Reference(ws, min_col=6, min_row=50, max_row=50 + daily_rows), titles_from_data=True)
        latency.set_categories(Reference(ws, min_col=2, min_row=51, max_row=50 + daily_rows))
        latency.height = 7
        latency.width = 13
        latency.style = 13
        latency.display_blanks = "gap"
        latency.legend = None
        latency.series[0].graphicalProperties.line.solidFill = FOUNDRY_TEAL
        latency.series[0].marker.symbol = "circle"
        latency.series[0].tx.strRef = None
        latency.series[0].tx.v = "Average latency"
        ws.add_chart(latency, "A32")

    if model_rows:
        models = BarChart()
        models.type = "bar"
        models.title = "Cost by model"
        models.x_axis.title = "Cost"
        models.add_data(Reference(ws, min_col=10, min_row=50, max_row=50 + model_rows), titles_from_data=True)
        models.set_categories(Reference(ws, min_col=9, min_row=51, max_row=50 + model_rows))
        models.height = 7
        models.width = 13
        models.legend = None
        models.style = 10
        models.series[0].graphicalProperties.solidFill = AZURE_BLUE
        ws.add_chart(models, "H32")


def _write_interactive_dashboard(
    workbook: Workbook,
    data: pd.DataFrame,
    *,
    minimum_date: dt.date,
    maximum_date: dt.date,
    currency: str,
) -> None:
    ws = workbook.create_sheet("Dashboard", 1)
    _style_dashboard(
        ws,
        "Multi-cloud LLM token observability",
        f"Derived view; token estimates and quality gaps are explicit. Cost currency: {currency}.",
        end_col=14,
    )
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 85
    _write_dashboard_filters(ws, minimum_date=minimum_date, maximum_date=maximum_date)
    _write_dashboard_kpis(ws, has_data=not data.empty)
    for cell, formula, color in (
        ("K8", "K8<1", "FFF4CE"),
        ("M8", "M8<1", "FFF4CE"),
        ("A12", "A12<1", "FFF4CE"),
        ("C12", "C12>0", "FFF4CE"),
        ("E12", "E12>0", "FFF4CE"),
        ("G12", "G12>0", "FFF4CE"),
        ("I12", "I12>0", "FDE7E9"),
        ("K12", "K12>0", "FDE7E9"),
        ("M12", "M12>0", "FDE7E9"),
    ):
        ws.conditional_formatting.add(cell, FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)))
    daily_rows, model_rows = _write_dashboard_helper_tables(ws, data)
    _add_interactive_dashboard_charts(ws, daily_rows, model_rows)
    for column in range(1, 15):
        ws.column_dimensions[get_column_letter(column)].width = 12
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[12].height = 25
    ws.print_area = "A1:N46"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.selection[0].activeCell = "B5"
    ws.sheet_view.selection[0].sqref = "B5"
    workbook.active = workbook.index(ws)


def _add_cost_charts(ws, day_rows: int, model_rows: int) -> None:
    if day_rows:
        chart = AreaChart()
        chart.title = "Cost by day"
        chart.y_axis.title = "Cost"
        chart.x_axis.title = "Date"
        chart.y_axis.numFmt = "0.000000"
        chart.add_data(Reference(ws, min_col=2, min_row=8, max_row=8 + day_rows), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8 + day_rows))
        chart.height = 7
        chart.width = 12
        chart.style = 13
        chart.legend = None
        chart.series[0].tx.strRef = None
        chart.series[0].tx.v = str(ws.cell(8, 2).value)
        ws.add_chart(chart, "A20")
    if model_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Cost by model"
        chart.x_axis.title = "Cost"
        chart.y_axis.title = "Model"
        chart.x_axis.numFmt = "0.000000"
        chart.add_data(Reference(ws, min_col=5, min_row=8, max_row=8 + model_rows), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=4, min_row=9, max_row=8 + model_rows))
        chart.height = 7
        chart.width = 12
        chart.style = 10
        chart.legend = None
        chart.series[0].tx.strRef = None
        chart.series[0].tx.v = str(ws.cell(8, 5).value)
        ws.add_chart(chart, "J20")


def _latency_wide(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=["date"])
    pivot = frame.pivot(index="date", columns="model", values=["average_latency_ms", "p95_latency_ms"])
    pivot.columns = [f"{model} {'avg' if metric == 'average_latency_ms' else 'p95'}" for metric, model in pivot.columns]
    return pivot.reset_index().sort_values("date")


def _add_tokens_latency_charts(ws, token_frame: pd.DataFrame, latency_frame: pd.DataFrame) -> None:
    if not token_frame.empty and len(token_frame.columns) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Contributing tokens by day and type (exact + estimate)"
        chart.y_axis.title = "Tokens"
        chart.add_data(
            Reference(ws, min_col=2, max_col=len(token_frame.columns), min_row=8, max_row=8 + len(token_frame)),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8 + len(token_frame)))
        chart.height = 7
        chart.width = 13
        chart.style = 12
        for series, title in zip(chart.series, token_frame.columns[1:], strict=True):
            series.tx.strRef = None
            series.tx.v = str(title)
        ws.add_chart(chart, "A20")
    if not latency_frame.empty and len(latency_frame.columns) > 1:
        start_col = 10
        chart = LineChart()
        chart.title = "Request latency by model and day"
        chart.y_axis.title = "Milliseconds"
        chart.add_data(
            Reference(
                ws,
                min_col=start_col + 1,
                max_col=start_col + len(latency_frame.columns) - 1,
                min_row=8,
                max_row=8 + len(latency_frame),
            ),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=start_col, min_row=9, max_row=8 + len(latency_frame)))
        chart.height = 7
        chart.width = 13
        chart.style = 13
        chart.display_blanks = "gap"
        for series, title in zip(chart.series, latency_frame.columns[1:], strict=True):
            series.tx.strRef = None
            series.tx.v = str(title)
            series.marker.symbol = "circle"
            series.marker.size = 6
        ws.add_chart(chart, "J20")


def _date_labels(frame: pd.DataFrame, column: str = "Date") -> pd.DataFrame:
    """Use ISO text labels on presentation sheets so charts never show Excel serials."""
    result = frame.copy()
    if column in result.columns:
        result[column] = result[column].map(
            lambda value: value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)
        )
    return result


def _autosize(ws, *, max_width: int = 28) -> None:
    for column_cells in ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)):
        width = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8) + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 10), max_width)


def _write_data_quality_sheet(workbook: Workbook, summaries: dict[str, Any], report: LoadReport) -> None:
    ws = workbook.create_sheet("Data Quality", 2)
    _style_dashboard(
        ws,
        "Data quality and provenance",
        "Runtime evidence only. Missing coverage is unknown, never silently promoted to zero or proven.",
        end_col=15,
    )
    quality = summaries["quality"]
    skipped = report.malformed_lines + report.schema_invalid_lines
    kpis = (
        ("A4", "A5", "Quality status", quality.get("quality_status", "clean"), "General"),
        ("C4", "C5", "Volume status", quality.get("volume_status", "ok"), "General"),
        ("E4", "E5", "Pricing coverage", quality.get("pricing_coverage"), "0.0%"),
        ("G4", "G5", "Instrumented latency", quality.get("instrumented_latency_coverage"), "0.0%"),
        ("I4", "I5", "Provider-total coverage", quality.get("provider_total_coverage"), "0.0%"),
        ("K4", "K5", "Coverage status", quality.get("coverage_status", "missing"), "General"),
        ("A7", "A8", "Known exact token share", quality.get("known_exact_token_share"), "0.0%"),
        ("C7", "C8", "Active flagged events", quality.get("active_flagged_event_count", 0), "#,##0"),
        ("E7", "E8", "Usage-loss events", quality.get("usage_loss_event_count", 0), "#,##0"),
        ("G7", "G8", "Unknown quantities", quality.get("unknown_quantity_count", 0), "#,##0"),
        ("I7", "I8", "Skipped / duplicate rows", skipped + report.duplicate_event_ids, "#,##0"),
        ("K7", "K8", "Latency applicability", quality.get("latency_applicability"), "0.0%"),
        ("M7", "M8", "Stream TTFT coverage", quality.get("ttft_coverage"), "0.0%"),
    )
    for label_cell, value_cell, label, value, number_format in kpis:
        _write_kpi(ws, label_cell, value_cell, label, value, number_format)

    status_cell = ws["A5"]
    status_cell.fill = PatternFill(
        "solid",
        fgColor={"clean": "DFF6DD", "warning": "FFF4CE", "blocked": "FDE7E9"}.get(
            str(status_cell.value), LIGHT
        ),
    )
    coverage_cell = ws["K5"]
    coverage_cell.fill = PatternFill(
        "solid",
        fgColor={"complete": "DFF6DD", "partial": "FFF4CE", "missing": "FDE7E9"}.get(
            str(coverage_cell.value), LIGHT
        ),
    )

    ws["A10"] = "Provider/runtime coverage"
    ws["A10"].font = Font(size=11, bold=True, color=NAVY)
    provider_columns = [
        "cloud_provider",
        "provider",
        "api_surface",
        "Events",
        "Requests",
        "Contributing tokens",
        "Exact tokens",
        "Estimated tokens",
        "Unverified tokens",
        "Unknown quantities",
        "Flagged events",
        "Usage-loss events",
        "Provider-total coverage",
        "Latency coverage",
        "Pricing coverage",
    ]
    provider_summary = summaries["provider_summary"].reindex(columns=provider_columns)
    provider_end, _ = _write_frame(
        ws,
        provider_summary,
        start_row=11,
        start_col=1,
        table_name="ProviderRuntimeQuality",
    )
    source_title_row = provider_end + 3
    ws.cell(source_title_row, 1, "Source provenance")
    ws.cell(source_title_row, 1).font = Font(size=11, bold=True, color=NAVY)
    source_columns = [
        "source_kind",
        "source_file",
        "Events",
        "Requests",
        "Contributing tokens",
        "Exact tokens",
        "Estimated tokens",
        "Unverified tokens",
        "Unknown quantities",
        "Flagged events",
        "Usage-loss events",
        "Latency coverage",
        "Pricing coverage",
    ]
    source_summary = summaries["source_summary"].reindex(columns=source_columns)
    source_start_row = source_title_row + 1
    _write_frame(
        ws,
        source_summary,
        start_row=source_start_row,
        start_col=1,
        table_name="SourceProvenance",
    )

    for frame, header_row in ((provider_summary, 11), (source_summary, source_start_row)):
        for header in ("Provider-total coverage", "Latency coverage", "Pricing coverage"):
            if header not in frame.columns:
                continue
            column = frame.columns.get_loc(header) + 1
            for row_index in range(header_row + 1, header_row + len(frame) + 1):
                ws.cell(row_index, column).number_format = "0.0%"
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    _autosize(ws, max_width=45)


def _write_provider_readiness_sheet(workbook: Workbook) -> None:
    fixture_records = realistic_fixture_records()
    surface_rows = build_provider_validation_matrix(fixture_records)
    capability_rows = build_capability_certification_matrix(
        fixture_records,
        PROVIDER_CAPABILITY_POLICIES,
    )
    capability_summary = summarize_capability_certification(capability_rows)
    surface_frame = pd.DataFrame(
        [
            {
                "Certification": row["certification_status"],
                "Provider": row["provider"],
                "API surface": row["api_surface"],
                "Adapter": row["adapter_name"],
                "Real fixtures": row["real_fixture_count"],
                "Simulated fixtures": row["simulated_fixture_count"],
                "Remaining gaps": ", ".join(row["gaps"]) or "none",
            }
            for row in surface_rows
        ]
    )
    def evidence_label(items: list[str]) -> str:
        if len(items) <= 2:
            return ", ".join(items) or "none"
        return f"{', '.join(items[:2])} (+{len(items) - 2} more)"

    capability_frame = pd.DataFrame(
        [
            {
                "Certification": row["certification_status"],
                "Provider": row["provider"],
                "API surface": row["api_surface"],
                "Capability": row["capability"],
                "Real fixtures": row["real_fixture_count"],
                "Simulated fixtures": row["simulated_fixture_count"],
                "Evidence": evidence_label(row["evidence"]),
                "Limitation": row["note"] or "none",
            }
            for row in capability_rows
        ]
    )

    ws = workbook.create_sheet("Provider Readiness", 3)
    _style_dashboard(
        ws,
        "Provider readiness",
        "Codebase fixture evidence. REAL proof is required per capability; shared code and simulations do not certify a cloud.",
        end_col=12,
    )
    kpis = (
        ("A4", "A5", "Proven capabilities", capability_summary["proven_count"]),
        ("C4", "C5", "Simulated capabilities", capability_summary["simulated_count"]),
        ("E4", "E5", "Unvalidated capabilities", capability_summary["unvalidated_count"]),
        ("G4", "G5", "Unsupported claims", capability_summary["unsupported_count"]),
    )
    for label_cell, value_cell, label, value in kpis:
        _write_kpi(ws, label_cell, value_cell, label, value, "#,##0")
    ws["A8"] = "Surface certification"
    ws["A8"].font = Font(size=11, bold=True, color=NAVY)
    surface_end, _ = _write_frame(
        ws,
        surface_frame,
        start_row=9,
        start_col=1,
        table_name="ProviderSurfaceCertification",
    )
    capability_title_row = surface_end + 3
    ws.cell(capability_title_row, 1, "Capability certification")
    ws.cell(capability_title_row, 1).font = Font(size=11, bold=True, color=NAVY)
    _write_frame(
        ws,
        capability_frame,
        start_row=capability_title_row + 1,
        start_col=1,
        table_name="ProviderCapabilityCertification",
    )
    status_fills = {
        "proven": "DFF6DD",
        "partially_proven": "FFF4CE",
        "simulated": "FFF4CE",
        "unvalidated": "FDE7E9",
        "unsupported": "E5E7EB",
    }
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            fill = status_fills.get(str(cell.value))
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(bold=True, color=NAVY)
    ws.freeze_panes = "A9"
    ws.sheet_view.showGridLines = False
    _autosize(ws, max_width=55)
    for column in ("G", "H"):
        ws.column_dimensions[column].width = 55
        for cell in ws[column]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_dashboard(
    data: pd.DataFrame,
    summaries: dict[str, Any],
    report: LoadReport,
    output_path: str | os.PathLike[str],
    *,
    max_data_rows: int = DEFAULT_MAX_DATA_ROWS,
) -> str:
    """Create auditable reporting sheets and native Excel chart objects from scratch."""
    if isinstance(max_data_rows, bool) or not isinstance(max_data_rows, int) or max_data_rows <= 0:
        raise ValueError("max_data_rows must be a positive integer")
    if len(data) + 1 > EXCEL_MAX_ROWS:
        raise ValueError(f"Data has {len(data)} rows and exceeds the Excel worksheet row limit")
    if len(data) > max_data_rows:
        raise ValueError(
            f"Data has {len(data)} quantity rows and exceeds the configured dashboard safety limit "
            f"of {max_data_rows}; use partitioned reporting or raise --max-data-rows deliberately"
        )
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.title = "LLM Token Observability Dashboard"
    workbook.properties.subject = "Derived reporting over append-only TokenEvent JSONL"

    ws_data = workbook.active
    ws_data.title = "Data"
    ws_data.sheet_view.showGridLines = False
    ws_data.freeze_panes = "A2"
    for row in dataframe_to_rows(data, index=False, header=True):
        ws_data.append([_excel_value(value) for value in row])
    for cell in ws_data[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_data.row_dimensions[1].height = 34
    if not data.empty:
        table = Table(displayName="DataTable", ref=f"A1:{get_column_letter(len(DATA_COLUMNS))}{len(data) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws_data.add_table(table)
        quality_column = DATA_COLUMNS.index("cost_quality") + 1
        quality_letter = get_column_letter(quality_column)
        ws_data.conditional_formatting.add(
            f"{quality_letter}2:{quality_letter}{len(data) + 1}",
            FormulaRule(
                formula=[f'OR({quality_letter}2="missing_price",{quality_letter}2="subtotal_exceeds_parent")'],
                fill=PatternFill("solid", fgColor="FEE2E2"),
            ),
        )
    for column_name in ("timestamp_utc", "date"):
        column_index = DATA_COLUMNS.index(column_name) + 1
        number_format = "yyyy-mm-dd hh:mm:ss" if column_name == "timestamp_utc" else "yyyy-mm-dd"
        for row_index in range(2, len(data) + 2):
            ws_data.cell(row_index, column_index).number_format = number_format
    for column_name in ("derived_cost", "unit_price_per_million"):
        column_index = DATA_COLUMNS.index(column_name) + 1
        for row_index in range(2, len(data) + 2):
            ws_data.cell(row_index, column_index).number_format = "0.000000"
    _autosize(ws_data, max_width=32)

    currency = summaries["currency"] or "currency unknown"
    ws_cost = workbook.create_sheet("Coûts")
    _style_dashboard(ws_cost, "Costs", f"Presentation-layer estimates in {currency}; missing prices remain blank, never zero.")
    _write_kpi(ws_cost, "A4", "A5", "Known total cost", summaries["total_cost"], "0.000000")
    _write_kpi(ws_cost, "C4", "C5", "Pricing coverage", summaries["pricing_coverage"], "0.0%")
    _write_kpi(ws_cost, "E4", "E5", "Valid events", report.valid_events, "#,##0")
    _write_kpi(ws_cost, "G4", "G5", "Skipped lines", report.malformed_lines + report.schema_invalid_lines, "#,##0")
    cost_day = _date_labels(
        summaries["cost_by_day"].rename(columns={"date": "Date", "derived_cost": f"Cost ({currency})"})
    )
    cost_model = summaries["cost_by_model"].rename(columns={"model": "Model", "derived_cost": f"Cost ({currency})"})
    _write_frame(ws_cost, cost_day, start_row=8, start_col=1, table_name="CostByDay")
    _write_frame(ws_cost, cost_model, start_row=8, start_col=4, table_name="CostByModel")
    _add_cost_charts(ws_cost, len(cost_day), len(cost_model))
    _autosize(ws_cost)

    ws_tokens = workbook.create_sheet("Tokens & Latence")
    _style_dashboard(
        ws_tokens,
        "Tokens and latency",
        "Token stacks use quantity_in_total; latency is measured once per request_correlation_id.",
    )
    contributing_tokens = int(data["event_contributing_tokens_once"].sum()) if not data.empty else 0
    request_count = int(data["request_count_once"].sum()) if not data.empty else 0
    _write_kpi(ws_tokens, "A4", "A5", "Contributing tokens", contributing_tokens, "#,##0")
    _write_kpi(ws_tokens, "C4", "C5", "Requests", request_count, "#,##0")
    request_latencies = (
        data.loc[data["request_count_once"] == 1, "request_latency_ms"].dropna()
        if not data.empty
        else pd.Series(dtype=float)
    )
    average_latency = float(request_latencies.mean()) if not request_latencies.empty else None
    p95_latency = float(request_latencies.quantile(0.95)) if not request_latencies.empty else None
    _write_kpi(ws_tokens, "E4", "E5", "Average latency (ms)", average_latency, "0.0")
    _write_kpi(ws_tokens, "G4", "G5", "P95 latency (ms)", p95_latency, "0.0")
    tokens_frame = _date_labels(summaries["tokens_by_day"].rename(columns={"date": "Date"}))
    latency_frame = _date_labels(
        _latency_wide(summaries["latency_by_day_model"]).rename(columns={"date": "Date"})
    )
    _write_frame(ws_tokens, tokens_frame, start_row=8, start_col=1, table_name="TokensByDay")
    _write_frame(ws_tokens, latency_frame, start_row=8, start_col=10, table_name="LatencyByDay")
    _add_tokens_latency_charts(ws_tokens, tokens_frame, latency_frame)
    _autosize(ws_tokens)

    ws_use = workbook.create_sheet("Use cases")
    _style_dashboard(ws_use, "Use cases", "Cost, contributing tokens and distinct requests by workflow/use case.")
    use_cases = summaries["use_cases"].rename(
        columns={
            "use_case": "Use case",
            "derived_cost": f"Cost ({currency})",
            "contributing_tokens": "Contributing tokens",
            "requests": "Requests",
        }
    )
    _write_frame(ws_use, use_cases, start_row=5, start_col=1, table_name="UseCaseSummary")
    if not use_cases.empty:
        pie_column = 2 if use_cases.iloc[:, 1].notna().any() else 3
        chart = PieChart()
        chart.title = "Cost share by use case" if pie_column == 2 else "Token share by use case"
        chart.add_data(Reference(ws_use, min_col=pie_column, min_row=5, max_row=5 + len(use_cases)), titles_from_data=True)
        chart.set_categories(Reference(ws_use, min_col=1, min_row=6, max_row=5 + len(use_cases)))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.separator = "\n"
        chart.legend = None
        chart.height = 8
        chart.width = 11
        chart.style = 10
        ws_use.add_chart(chart, "F5")
    _autosize(ws_use)

    ws_lists = workbook.create_sheet("_Lists")
    minimum_date, maximum_date = _write_dashboard_lists(workbook, ws_lists, data)
    _write_interactive_dashboard(
        workbook,
        data,
        minimum_date=minimum_date,
        maximum_date=maximum_date,
        currency=currency,
    )
    _write_data_quality_sheet(workbook, summaries, report)
    _write_provider_readiness_sheet(workbook)

    target = Path(output_path).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.{os.getpid()}.tmp")
    workbook.save(temporary)
    os.replace(temporary, target)
    return str(target)


def generate_dashboard(
    *,
    data_dir: str | os.PathLike[str],
    prices_path: str | os.PathLike[str] | None,
    output_path: str | os.PathLike[str],
    recursive: bool = False,
    max_data_rows: int = DEFAULT_MAX_DATA_ROWS,
) -> tuple[str, LoadReport, dict[str, Any]]:
    events, report = load_jsonl_events(data_dir, recursive=recursive)
    prices = load_prices(prices_path)
    data = build_data_frame(events, prices)
    summaries = build_summary_frames(data)
    path = write_dashboard(data, summaries, report, output_path, max_data_rows=max_data_rows)
    return path, report, summaries


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Generate a native Excel token dashboard from TokenEvent JSONL")
    parser.add_argument("--data-dir", default="data")
    parser.add_argument("--prices", default="prices.csv")
    parser.add_argument("--output", default="dashboard.xlsx")
    parser.add_argument("--recursive", action="store_true", help="include JSONL files in nested partition directories")
    parser.add_argument(
        "--max-data-rows",
        type=int,
        default=DEFAULT_MAX_DATA_ROWS,
        help="fail before creating an oversized workbook (quantity-grain rows)",
    )
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--log-level", default="INFO", choices=("DEBUG", "INFO", "WARNING", "ERROR"))
    args = parser.parse_args(argv)
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(levelname)s %(message)s")
    output, report, summaries = generate_dashboard(
        data_dir=args.data_dir,
        prices_path=args.prices,
        output_path=args.output,
        recursive=args.recursive,
        max_data_rows=args.max_data_rows,
    )
    result = {
        "output": output,
        "files_read": report.files_read,
        "lines_read": report.lines_read,
        "valid_events": report.valid_events,
        "skipped_lines": report.malformed_lines + report.schema_invalid_lines,
        "duplicate_event_ids": report.duplicate_event_ids,
        "currency": summaries["currency"],
        "total_cost": summaries["total_cost"],
        "pricing_coverage": summaries["pricing_coverage"],
        "latency_coverage": summaries["quality"].get("latency_coverage"),
        "instrumented_latency_coverage": summaries["quality"].get("instrumented_latency_coverage"),
        "latency_applicability": summaries["quality"].get("latency_applicability"),
        "provider_total_coverage": summaries["quality"].get("provider_total_coverage"),
        "data_row_count": summaries["quality"].get("data_row_count", 0),
        "volume_status": summaries["quality"].get("volume_status", "ok"),
        "quality_status": summaries["quality"].get("quality_status", "clean"),
        "coverage_status": summaries["quality"].get("coverage_status", "missing"),
    }
    if args.json:
        print(json.dumps(result, ensure_ascii=True, sort_keys=True))
    else:
        print(f"dashboard: {output}")
        print(
            f"events={report.valid_events} skipped={result['skipped_lines']} "
            f"pricing_coverage={summaries['pricing_coverage']}"
        )


if __name__ == "__main__":
    main()
