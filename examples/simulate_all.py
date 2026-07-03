"""Full simulated run — exercise EVERY adapter + streaming + RAG/agent through the real
pipeline, deliver over HTTP, persist, export to Excel, and reconcile every total.

Run: & "C:\\Users\\yerabhaoui\\python-portable\\python.exe" examples\\simulate_all.py

No API credit: each "call" replays a SIMULATED provider fixture, but every layer it flows
through is the real thing (façade -> collector -> live HTTP server -> JSONL -> Excel). At the
end it checks that the model total == the server /v1/stats total == the exported Excel total.
"""

import json
import os
import sys
import tempfile
import threading
from urllib import request as urlreq

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
FIX = os.path.join(ROOT, "tests", "fixtures")
OUT = os.path.join(ROOT, "demo_output")
sys.path.insert(0, ROOT)

import openpyxl  # noqa: E402

from api.main import create_server, make_http_transport  # noqa: E402
from tracker.adapters.anthropic_messages_adapter import AnthropicMessagesAdapter  # noqa: E402
from tracker.adapters.azure_openai_chat_completions_adapter import AzureOpenAIChatCompletionsAdapter  # noqa: E402
from tracker.adapters.bedrock_converse_adapter import BedrockConverseAdapter  # noqa: E402
from tracker.adapters.bedrock_invoke_model_adapter import BedrockInvokeModelAdapter  # noqa: E402
from tracker.adapters.gemini_generate_content_adapter import GeminiGenerateContentAdapter  # noqa: E402
from tracker.adapters.openai_chat_completions_adapter import OpenAIChatCompletionsAdapter  # noqa: E402
from tracker.adapters.openai_responses_adapter import OpenAIResponsesAdapter  # noqa: E402
from tracker.analytics.coverage import build_coverage_exactness  # noqa: E402
from tracker.collector.client import CollectorClient  # noqa: E402
from tracker.context.propagation import span, trace  # noqa: E402
from tracker.derive.trace_rollup import roll_up  # noqa: E402
from tracker.export.excel_exporter import export_excel  # noqa: E402
from tracker.models.trace import Trace  # noqa: E402
from tracker.service import track_response, track_stream  # noqa: E402
from tracker.storage.file_repository import FileRepository  # noqa: E402
from tracker.workflows.agent_tracker import new_tool_span, record_tool_result  # noqa: E402
from tracker.workflows.rag_tracker import new_rag_span, record_retrieved_context, record_vector_search  # noqa: E402

checks = []


def ok(cond, label):
    checks.append(bool(cond))
    print(("  [OK] " if cond else "  [XX] ") + label)


def load(name):
    with open(os.path.join(FIX, name), encoding="utf-8") as f:
        return json.load(f)["response"]


def get(base, path):
    with urlreq.urlopen(base + path, timeout=5) as r:
        return json.loads(r.read())


CALLS = [
    ("OpenAI Chat", OpenAIChatCompletionsAdapter(), "openai_chat_completions_cached_reasoning.SIMULATED.json", 1300),
    ("OpenAI Responses", OpenAIResponsesAdapter(), "openai_responses_cached_reasoning.SIMULATED.json", 1300),
    ("Azure OpenAI Chat", AzureOpenAIChatCompletionsAdapter(), "openai_chat_completions_cached_reasoning.SIMULATED.json", 1300),
    ("Bedrock Converse", BedrockConverseAdapter(), "bedrock_converse_cache.SIMULATED.json", 1300),
    ("Gemini", GeminiGenerateContentAdapter(), "gemini_generate_content_thinking.SIMULATED.json", 1550),
    ("Anthropic", AnthropicMessagesAdapter(), "anthropic_messages_cache.SIMULATED.json", 1300),
    ("Bedrock InvokeModel", BedrockInvokeModelAdapter(), "bedrock_invoke_model_headers.SIMULATED.json", 1300),
]


def main():
    work = tempfile.mkdtemp(prefix="tt_simall_")
    store = os.path.join(work, "events.jsonl")
    repo = FileRepository(store)
    server = create_server(repo, "127.0.0.1", 0)
    base = f"http://127.0.0.1:{server.server_address[1]}"
    threading.Thread(target=server.serve_forever, daemon=True).start()
    collector = CollectorClient(make_http_transport(base + "/v1/events"))

    print(f"\nCollector server up at {base}\n")
    print(f"  {'call':<22}{'provider':<14}{'surface':<18}{'tokens':>8}  flags")
    print("  " + "-" * 74)

    with trace(business_id="acme", workflow="multi_provider", environment="prod") as root:
        tr = Trace(trace_id=root.trace_id, business_id=root.business_id, workflow=root.workflow, environment=root.environment)

        # 1) one simulated call per adapter, through the façade + collector
        for label, adapter, fixture, expected in CALLS:
            with span():
                res = track_response(load(fixture), adapter, trace=tr, collector=collector)
            ev = res.event
            ok(ev.event_contributing_tokens == expected, f"{label}: contributes {expected}")
            flags = ",".join(ev.data_quality_flags) or "-"
            print(f"  {label:<22}{ev.provider:<14}{ev.api_surface:<18}{ev.event_contributing_tokens:>8}  {flags}")

        # 2) a streamed call: interrupted (partial estimate) then resolved (supersedes it)
        with span() as sctx:
            st = track_stream(context=sctx, provider="openai", api_surface="chat_completions", model="gpt-4o")
            st.feed("Streaming the answer ")
            st.feed("but the connection dropped")
            partial = st.interrupt()
            final = st.resolve_with_final_usage(output_tokens=120, input_tokens=300, provider_total_tokens=420)
            for e in (partial, final):
                collector.record(e)
                tr.add_event(e)
        ok(partial.superseded and partial.event_contributing_tokens == 0, "stream partial superseded -> contributes 0")
        ok(final.event_contributing_tokens == 420, "stream final contributes 420")
        print(
            f"  {'Stream (interrupted)':<22}{'openai':<14}{'chat_completions':<18}{final.event_contributing_tokens:>8}  partial superseded"
        )

        # 3) RAG + agent spans (annotations — must NOT change the token total)
        total_before_spans = roll_up(tr).observed_total_contributing_tokens
        with span() as vs:
            v = new_rag_span(root.trace_id, "vector_search", parent_span_id=root.span_id, span_id=vs.span_id)
            record_vector_search(v, num_results=5, latency_ms=7.1, index="faiss")
            tr.add_span(v)
        with span() as pa:
            p = new_rag_span(root.trace_id, "prompt_assembly", parent_span_id=root.span_id, span_id=pa.span_id)
            record_retrieved_context(p, context_text="retrieved policy text " * 15, injected_into_prompt=True)
            tr.add_span(p)
        with span() as tl:
            t = new_tool_span(root.trace_id, tool_name="search_orders", tool_call_id="c1", parent_span_id=root.span_id, span_id=tl.span_id)
            record_tool_result(t, result_text="orders found " * 10, injected_into_context=True)
            tr.add_span(t)
        ok(
            roll_up(tr).observed_total_contributing_tokens == total_before_spans,
            "RAG/agent spans do not change the token total (no double count)",
        )

    # 4) flush everything to the live server over HTTP
    while collector.pending:
        collector.flush()
    ok(collector.sent_total == len(CALLS) + 2, f"all {len(CALLS) + 2} events delivered over HTTP")
    ok(len(repo.read_all()) == len(CALLS) + 2, "every event persisted to JSONL")

    # 5) reconcile model vs server vs Excel
    model_total = roll_up(tr).observed_total_contributing_tokens
    server_total = get(base, "/v1/stats")["total"]
    xlsx = os.path.join(OUT, "simulate_all.xlsx")
    os.makedirs(OUT, exist_ok=True)
    export_excel(tr, xlsx)
    wb = openpyxl.load_workbook(xlsx)
    rows = list(wb["TokenEvents"].iter_rows(values_only=True))
    idx = rows[0].index("event_contributing_tokens")
    excel_total = sum(int(r[idx]) for r in rows[1:])

    cov = build_coverage_exactness(tr)
    print("\n  " + "-" * 74)
    print(f"  GRAND TOTAL : model={model_total}  server={server_total}  excel={excel_total}")
    print(
        f"  events={cov['event_count']}  superseded={cov['superseded_event_count']}  "
        f"spans={len(tr.spans)}  exactness={cov['exactness_ratio']}"
    )
    server.shutdown()

    ok(model_total == server_total == excel_total, f"all three totals reconcile ({model_total})")
    ok("TokenSpans" in wb.sheetnames and len(list(wb["TokenSpans"].iter_rows())) - 1 == 3, "3 spans exported to Excel")

    passed = sum(checks)
    print(f"\n==> SIMULATION : {passed}/{len(checks)} verifications OK")
    print(f"    excel : {xlsx}")
    return 0 if passed == len(checks) else 1


if __name__ == "__main__":
    sys.exit(main())
