"""Operational smoke test — is the whole system actually working end to end?

Run: & "C:\\Users\\yerabhaoui\\python-portable\\python.exe" examples\\verify_operational.py

Not a unit test: it starts the REAL collector HTTP server, instruments two provider calls
through the real pipeline, delivers them over HTTP, and checks they are persisted, totalled,
and exportable to a real Excel that reopens correctly. Prints a final OPERATIONAL verdict.
"""

import json
import os
import sys
import tempfile
import threading

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
FIX = os.path.join(ROOT, "tests", "fixtures")
sys.path.insert(0, ROOT)

from urllib import request as urlreq  # noqa: E402

import openpyxl  # noqa: E402

from api.main import create_server, make_http_transport  # noqa: E402
from tracker.adapters.anthropic_messages_adapter import AnthropicMessagesAdapter  # noqa: E402
from tracker.adapters.openai_chat_completions_adapter import OpenAIChatCompletionsAdapter  # noqa: E402
from tracker.context.propagation import span, trace  # noqa: E402
from tracker.derive.trace_rollup import observed_total_contributing_tokens  # noqa: E402
from tracker.export.excel_exporter import export_excel  # noqa: E402
from tracker.models.trace import Trace  # noqa: E402
from tracker.normalization.normalizer import normalize  # noqa: E402
from tracker.storage.file_repository import FileRepository  # noqa: E402

checks = []


def ok(cond, label):
    checks.append(bool(cond))
    print(("  [OK] " if cond else "  [XX] ") + label)


def load(name):
    with open(os.path.join(FIX, name), encoding="utf-8") as f:
        return json.load(f)["response"]


def get(base, path):
    with urlreq.urlopen(base + path, timeout=5) as r:
        return r.status, json.loads(r.read())


def main():
    work = tempfile.mkdtemp(prefix="tt_ops_")
    store = os.path.join(work, "collector_store.jsonl")
    xlsx = os.path.join(work, "operational_tokens.xlsx")

    # 1) start the REAL collector server on an ephemeral loopback port
    repo = FileRepository(store)
    server = create_server(repo, "127.0.0.1", 0)
    base = f"http://127.0.0.1:{server.server_address[1]}"
    threading.Thread(target=server.serve_forever, daemon=True).start()
    print(f"\nCollector server up at {base}")

    try:
        code, body = get(base, "/healthz")
        ok(code == 200 and body.get("status") == "ok", "le serveur collecteur repond (/healthz)")

        # 2) instrument two real provider calls through the full pipeline
        events = []
        with trace(business_id="ops-check", workflow="smoke", environment="prod"):
            with span():
                events.append(normalize(load("openai_chat_completions_cached_reasoning.SIMULATED.json"), OpenAIChatCompletionsAdapter()))
            with span():
                events.append(normalize(load("anthropic_messages_cache.SIMULATED.json"), AnthropicMessagesAdapter()))
        expected = sum(e.event_contributing_tokens for e in events)

        # 3) deliver them to the running server over HTTP
        transport = make_http_transport(base + "/v1/events")
        acked = transport([e.to_dict() for e in events])
        ok(len(acked) == 2, f"2 evenements livres en HTTP (acked={acked})")

        # 4) persisted to JSONL on disk
        back = repo.read_all()
        ok(len(back) == 2 and os.path.getsize(store) > 0, f"persistes en JSONL ({os.path.getsize(store)} octets sur disque)")

        # 5) totals reconcile via the live /v1/stats endpoint
        code, stats = get(base, "/v1/stats")
        ok(stats["total"] == expected, f"/v1/stats total reconcilie ({stats['total']} == {expected})")

        # 6) export a real Excel and reopen it
        tr = Trace(trace_id=back[0].trace_id)
        for e in back:
            tr.add_event(e)
        export_excel(tr, xlsx)
        wb = openpyxl.load_workbook(xlsx)
        rows = list(wb["TokenEvents"].iter_rows(values_only=True))
        idx = rows[0].index("event_contributing_tokens")
        xlsx_total = sum(int(r[idx]) for r in rows[1:])
        ok(os.path.exists(xlsx) and "TokenEvents" in wb.sheetnames, f"Excel genere et relisible ({len(wb.sheetnames)} feuilles)")
        ok(xlsx_total == observed_total_contributing_tokens(tr) == expected, f"total Excel == modele == {expected}")
    finally:
        server.shutdown()

    passed = sum(checks)
    print(f"\n==> OPERATIONNEL : {passed}/{len(checks)} verifications OK")
    print(f"    store : {store}")
    print(f"    excel : {xlsx}")
    return 0 if passed == len(checks) else 1


if __name__ == "__main__":
    sys.exit(main())
