"""Phase 9 — Excel export materializes the same totals as the model.

Run: python tests/test_csv_excel_export.py

Writes a real .xlsx (openpyxl), reopens it, and asserts the TokenEvents / TokenQuantities /
CoverageExactness sheets all carry the same trace total as the in-memory model — and that a
superseded event's contributing cell is 0.
"""

import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from tracker.derive.trace_rollup import observed_total_contributing_tokens  # noqa: E402
from tracker.export.excel_exporter import export_excel  # noqa: E402
from tracker.models.enums import PrecisionLevel, TokenType, UsageSource  # noqa: E402
from tracker.models.token_event import TokenEvent  # noqa: E402
from tracker.models.token_quantity import TokenQuantity  # noqa: E402
from tracker.models.trace import Trace  # noqa: E402
from tracker.normalization.additivity import assign_additivity  # noqa: E402

_failures = 0


def check(cond: bool, msg: str) -> None:
    global _failures
    if cond:
        print(f"[PASS] {msg}")
    else:
        _failures += 1
        print(f"[FAIL] {msg}")


def q(tt, qty, src=UsageSource.PROVIDER_RESPONSE, precision=PrecisionLevel.EXACT):
    additivity, subtotal_of = assign_additivity("openai", "responses", tt)
    return TokenQuantity(
        token_type=tt,
        quantity=qty,
        precision_level=precision,
        usage_source=src,
        additivity=additivity,
        subtotal_of=subtotal_of,
    )


a = TokenEvent(
    event_id="evt-a",
    request_correlation_id="r-a",
    trace_id="t-1",
    span_id="s-1",
    provider="openai",
    api_surface="responses",
    quantities=[q(TokenType.INPUT, 1000), q(TokenType.OUTPUT, 300), q(TokenType.CACHED_INPUT, 800), q(TokenType.REASONING, 250)],
    provider_total_tokens=1300,
    observation={"authoritative": True},
)
b = TokenEvent(
    event_id="evt-b",
    request_correlation_id="r-b",
    trace_id="t-1",
    span_id="s-2",
    provider="openai",
    api_surface="responses",
    quantities=[q(TokenType.OUTPUT, 200)],
    provider_total_tokens=200,
    observation={"authoritative": True},
)
c = TokenEvent(
    event_id="evt-c",
    request_correlation_id="r-b",
    trace_id="t-1",
    span_id="s-2",
    provider="openai",
    api_surface="responses",
    quantities=[q(TokenType.OUTPUT, 40, UsageSource.PARTIAL_STREAM_TOKENIZER, PrecisionLevel.ESTIMATE)],
    superseded=True,
    superseded_by="evt-b",
    data_quality_flags=["superseded"],
    observation={"authoritative": True},
)

trace = Trace(trace_id="t-1")
for e in (a, b, c):
    trace.add_event(e)
model_total = observed_total_contributing_tokens(trace)

out_dir = os.path.join(os.getcwd(), ".test_xlsx_out")
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, "tokens.xlsx")
export_excel(trace, out)
check(os.path.exists(out), "an .xlsx file was written")

wb = openpyxl.load_workbook(out)
for sheet in ("TokenQuantities", "TokenEvents", "TokenSpans", "CoverageExactness"):
    check(sheet in wb.sheetnames, f"sheet '{sheet}' is present")


def rows(ws):
    data = list(ws.iter_rows(values_only=True))
    header = list(data[0])
    return header, [dict(zip(header, r, strict=True)) for r in data[1:]]


# --- TokenEvents: sum event_contributing_tokens ---
eh, erows = rows(wb["TokenEvents"])
egrain = sum(int(r["event_contributing_tokens"]) for r in erows)
check(egrain == model_total, f"Excel TokenEvents sum == model (got {egrain})")
sup = next(r for r in erows if r["event_id"] == "evt-c")
check(int(sup["event_contributing_tokens"]) == 0, "Excel: superseded event contributes 0")

# --- TokenQuantities: sum quantity_in_total over non-superseded ---
qh, qrows = rows(wb["TokenQuantities"])
qgrain = sum(int(r["quantity_in_total"]) for r in qrows if str(r["event_superseded"]) == "False")
check(qgrain == model_total, f"Excel TokenQuantities sum (non-superseded) == model (got {qgrain})")

# --- CoverageExactness: the observed total cell ---
ch, crows = rows(wb["CoverageExactness"])
cov_row = next(r for r in crows if r.get("metric") == "observed_total_contributing_tokens")
check(int(cov_row["value"]) == model_total, "Excel CoverageExactness observed total == model")

check(egrain == qgrain == model_total, "Excel event-grain and quantity-grain agree with the model")

wb.close()
shutil.rmtree(out_dir, ignore_errors=True)

print("\nRESULT:", "all checks passed" if _failures == 0 else f"{_failures} FAILURE(S)")
sys.exit(1 if _failures else 0)
