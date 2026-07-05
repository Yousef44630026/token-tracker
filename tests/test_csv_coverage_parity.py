"""Regression — CSV/Excel parity for the headline CoverageExactness metrics.

The CoverageExactness metrics (observed_total_contributing_tokens, total_is_lower_bound,
unverified_quantity_count, exactness ratios...) are the project's headline numbers. They were
exported as an Excel sheet but export_csv never wrote a coverage_exactness.csv — so a
CSV/Power BI consumer got every secondary metric file (latency, reliability, cache...) but not
the one that says what the total IS and whether it is a floor. Both exports must carry it,
with identical values.

Run: python tests/test_csv_coverage_parity.py
"""

import csv
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tracker.analytics.coverage import build_coverage_exactness  # noqa: E402
from tracker.export.csv_exporter import export_csv  # noqa: E402
from tracker.models.enums import Additivity, PrecisionLevel, TokenType, UsageSource  # noqa: E402
from tracker.models.token_event import TokenEvent  # noqa: E402
from tracker.models.token_quantity import TokenQuantity  # noqa: E402
from tracker.models.trace import Trace  # noqa: E402

_failures = 0


def check(cond, msg):
    global _failures
    print(f"[{'PASS' if cond else 'FAIL'}] {msg}")
    if not cond:
        _failures += 1


# a trace whose headline status matters: one clean event + one unverified quantity -> floor
trace = Trace(trace_id="t")
trace.add_event(
    TokenEvent(
        event_id="e1",
        request_correlation_id="r1",
        trace_id="t",
        span_id="s",
        quantities=[
            TokenQuantity(TokenType.INPUT, 1000, PrecisionLevel.EXACT, UsageSource.PROVIDER_RESPONSE, Additivity.TOTAL_CONTRIBUTING),
            TokenQuantity(TokenType.OUTPUT, 500, PrecisionLevel.EXACT, UsageSource.PROVIDER_RESPONSE, Additivity.TOTAL_CONTRIBUTING),
        ],
        provider_total_tokens=1500,
    )
)
trace.add_event(
    TokenEvent(
        event_id="e2",
        request_correlation_id="r2",
        trace_id="t",
        span_id="s",
        quantities=[
            TokenQuantity(TokenType.CACHED_INPUT, 900, PrecisionLevel.EXACT, UsageSource.PROVIDER_RESPONSE, Additivity.UNVERIFIED),
        ],
    )
)

with tempfile.TemporaryDirectory() as out_dir:
    paths = export_csv(trace, out_dir)

    check("coverage_exactness" in paths, "export_csv returns a coverage_exactness path")
    cov_path = paths.get("coverage_exactness", os.path.join(out_dir, "coverage_exactness.csv"))
    check(os.path.exists(cov_path), "coverage_exactness.csv exists on disk")

    got: dict[str, str] = {}
    if os.path.exists(cov_path):
        with open(cov_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                got[row["metric"]] = row["value"]

    expected = build_coverage_exactness(trace)
    check(
        got.get("observed_total_contributing_tokens") == str(expected["observed_total_contributing_tokens"]),
        f"CSV observed total == model ({got.get('observed_total_contributing_tokens')} == {expected['observed_total_contributing_tokens']})",
    )
    check(expected["observed_total_contributing_tokens"] == 1500, "model total is 1500 (unverified 900 excluded)")
    check(got.get("total_is_lower_bound") == "True", "CSV carries total_is_lower_bound=True (the floor status travels)")
    check(got.get("unverified_quantity_count") == "1", "CSV carries unverified_quantity_count=1")
    check("exactness_ratio" in got, "CSV carries exactness_ratio")

    # every metric key present, none silently dropped
    missing = [k for k in expected if str(k) not in got]
    check(not missing, f"every CoverageExactness metric is in the CSV (missing: {missing})")

# Excel must still carry the same sheet with the same values (parity, not migration)
import openpyxl  # noqa: E402

from tracker.export.excel_exporter import export_excel  # noqa: E402

with tempfile.TemporaryDirectory() as out_dir:
    xlsx = export_excel(trace, os.path.join(out_dir, "t.xlsx"))
    wb = openpyxl.load_workbook(xlsx)
    check("CoverageExactness" in wb.sheetnames, "Excel still has the CoverageExactness sheet")
    ws = wb["CoverageExactness"]
    sheet = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
    check(
        sheet.get("observed_total_contributing_tokens") == 1500,
        "Excel CoverageExactness observed total still == model (1500)",
    )
    check(sheet.get("total_is_lower_bound") is True, "Excel CoverageExactness carries total_is_lower_bound")

print("\nRESULT:", "all checks passed" if _failures == 0 else f"{_failures} FAILURE(S)")
sys.exit(1 if _failures else 0)
