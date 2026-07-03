"""Derived operational metrics: latency, reliability, cache, RAG, agent, attribution."""

import csv
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from tracker.analytics.agent import build_agent_summary  # noqa: E402
from tracker.analytics.cache import build_cache_summary  # noqa: E402
from tracker.analytics.latency import build_latency_summary  # noqa: E402
from tracker.analytics.rag import build_rag_summary  # noqa: E402
from tracker.analytics.reliability import build_reliability_summary  # noqa: E402
from tracker.analytics.service_attribution import build_service_attribution  # noqa: E402
from tracker.export.csv_exporter import export_csv  # noqa: E402
from tracker.export.excel_exporter import export_excel  # noqa: E402
from tracker.models.enums import Additivity, PrecisionLevel, TokenType, UsageSource  # noqa: E402
from tracker.models.span import Span  # noqa: E402
from tracker.models.token_event import TokenEvent  # noqa: E402
from tracker.models.token_quantity import TokenQuantity  # noqa: E402
from tracker.models.trace import Trace  # noqa: E402

_failures = 0


def check(condition, message):
    global _failures
    print(f"[{'PASS' if condition else 'FAIL'}] {message}")
    if not condition:
        _failures += 1


def q(token_type, quantity, additivity=Additivity.TOTAL_CONTRIBUTING, *, subtotal_of=None, metadata=None):
    return TokenQuantity(
        token_type=token_type,
        quantity=quantity,
        precision_level=PrecisionLevel.EXACT,
        usage_source=UsageSource.PROVIDER_RESPONSE,
        additivity=additivity,
        subtotal_of=subtotal_of,
        metadata=metadata or {},
    )


trace = Trace(trace_id="trace-metrics", workflow="support", environment="prod")
trace.add_span(
    Span(
        span_id="vector-1",
        trace_id=trace.trace_id,
        span_type="vector_search",
        metadata={"num_results": 8, "latency_ms": 12.5},
    )
)
trace.add_span(
    Span(
        span_id="prompt-1",
        trace_id=trace.trace_id,
        span_type="prompt_assembly",
        metadata={
            "retrieved_context_estimated_tokens": 300,
            "retrieved_context_injected_into_prompt": True,
            "downstream_llm_span_id": "llm-1",
            "retrieved_context_hash": "ctx-a",
        },
    )
)
trace.add_span(Span(span_id="llm-1", trace_id=trace.trace_id, span_type="final_generation"))
trace.add_span(
    Span(
        span_id="agent-step-1",
        trace_id=trace.trace_id,
        span_type="agent_step",
        metadata={
            "agent_run_id": "run-1",
            "step_index": 1,
            "step_type": "reason",
            "retry_count": 1,
            "memory_read_count": 2,
            "memory_write_count": 1,
            "loop_count": 2,
            "max_steps_reached": False,
        },
    )
)
trace.add_span(
    Span(
        span_id="tool-1",
        trace_id=trace.trace_id,
        span_type="tool",
        metadata={
            "tool_name": "search_cases",
            "tool_call_id": "call-1",
            "latency_ms": 42,
            "tool_result_estimated_tokens": 120,
            "tool_result_injected_into_context": True,
        },
    )
)

trace.add_event(
    TokenEvent(
        event_id="azure-1",
        request_correlation_id="req-azure-1",
        trace_id=trace.trace_id,
        span_id="llm-1",
        workflow="support",
        environment="prod",
        provider="azure_openai",
        model="gpt-test",
        api_surface="responses",
        quantities=[
            q(TokenType.INPUT, 1000, metadata={"azure_deployment": "prod-gpt"}),
            q(TokenType.CACHED_INPUT, 200, Additivity.SUBTOTAL_OF, subtotal_of="input", metadata={"azure_deployment": "prod-gpt"}),
            q(TokenType.OUTPUT, 300, metadata={"azure_deployment": "prod-gpt"}),
            q(TokenType.REASONING, 50, Additivity.SUBTOTAL_OF, subtotal_of="output", metadata={"azure_deployment": "prod-gpt"}),
        ],
        provider_total_tokens=1300,
        observation={
            "status": "complete",
            "duration_ms": 1000,
            "time_to_first_token_ms": 120,
            "provider_request_id": "az-req",
            "provider_response_id": "az-resp",
            "service_name": "support-api",
            "tenant_id": "tenant-a",
            "cloud_provider": "azure",
            "region": "francecentral",
        },
    )
)
trace.add_event(
    TokenEvent(
        event_id="agent-llm",
        request_correlation_id="req-agent-llm",
        trace_id=trace.trace_id,
        span_id="agent-step-1",
        provider="openai",
        model="gpt-agent",
        api_surface="responses",
        quantities=[q(TokenType.INPUT, 100), q(TokenType.OUTPUT, 50)],
        provider_total_tokens=150,
        observation={"status": "complete", "duration_ms": 500},
    )
)
trace.add_event(
    TokenEvent(
        event_id="bedrock-error",
        request_correlation_id="req-bedrock-error",
        trace_id=trace.trace_id,
        span_id="tool-1",
        provider="bedrock",
        model="nova",
        api_surface="converse",
        quantities=[],
        data_quality_flags=["raw_usage_missing"],
        observation={
            "status": "rate_limited",
            "http_status": 429,
            "provider_error_code": "rate_limit",
            "retry_count": 2,
            "authoritative": False,
        },
    )
)

latency = build_latency_summary(trace)
check(latency["event_count"] == 2, "latency excludes non-authoritative error event")
check(latency["average_duration_ms"] == 750.0, "average duration is derived")
check(latency["p95_duration_ms"] == 1000.0, "p95 duration uses nearest-rank percentile")
check(latency["average_time_to_first_token_ms"] == 120.0, "TTFT is reported when present")

reliability = build_reliability_summary(trace)
check(reliability["event_count"] == 3, "reliability sees all events")
check(reliability["error_count"] == 1, "rate-limited event counts as error")
check(reliability["rate_limit_count"] == 1, "rate limit count is derived")
check(reliability["retry_count"] == 2, "retry count is summed")
check(reliability["missing_usage_count"] == 1, "missing usage flag is counted")
check(reliability["success_rate"] == 0.666667, "success rate is derived")

cache = build_cache_summary(trace)
check(cache["cache_read_tokens"] == 200, "cache read tokens are visible")
check(cache["cache_hit_rate"] == 0.181818, "cache hit rate uses prompt input denominator")
check(cache["cache_savings_tokens"] == 200, "cache summary carries token savings")

rag = build_rag_summary(trace)
check(rag["vector_search_count"] == 1, "RAG vector search span is counted")
check(rag["vector_search_results"] == 8, "RAG native result count is summed")
check(rag["average_vector_search_latency_ms"] == 12.5, "RAG vector latency is averaged")
check(rag["context_utilization_ratio"] == 0.3, "RAG context utilization links to downstream LLM input")

agent = build_agent_summary(trace)
check(agent["agent_run_count"] == 1, "agent run count is derived")
check(agent["tool_call_count"] == 1, "tool call count is derived")
check(agent["retry_count"] == 1, "agent retry count is derived from span metadata")
check(agent["tool_result_estimated_tokens"] == 120, "tool result estimate is visible but not token-total source")
check(agent["agent_contributing_tokens"] == 150, "agent token total uses events attached to agent spans")

attribution = build_service_attribution(trace)
azure_row = next(row for row in attribution["rows"] if row["provider"] == "azure_openai")
check(azure_row["service_name"] == "support-api", "service attribution keeps service name")
check(azure_row["cloud_provider"] == "azure", "service attribution keeps cloud provider")
check(azure_row["region"] == "francecentral", "service attribution keeps region")
check(azure_row["deployment"] == "prod-gpt", "service attribution reads Azure deployment metadata")
check(azure_row["contributing_tokens"] == 1300, "service attribution rolls up tokens by group")

out_dir = os.path.join(os.getcwd(), ".test_metrics_out")
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir, exist_ok=True)
csv_paths = export_csv(trace, out_dir)
for key in (
    "latency_summary",
    "reliability_summary",
    "observation_contract",
    "cache_efficiency",
    "rag_efficiency",
    "agent_efficiency",
    "service_attribution",
):
    check(key in csv_paths and os.path.exists(csv_paths[key]), f"CSV export writes {key}")
check("cost_summary" not in csv_paths, "CSV export does not write pricing/cost summary")

with open(csv_paths["service_attribution"], encoding="utf-8", newline="") as handle:
    service_rows = list(csv.DictReader(handle))
check(any(row["provider"] == "azure_openai" for row in service_rows), "service attribution CSV has Azure row")

xlsx = os.path.join(out_dir, "metrics.xlsx")
export_excel(trace, xlsx)
wb = openpyxl.load_workbook(xlsx)
for sheet in (
    "LatencySummary",
    "ReliabilitySummary",
    "ObservationContract",
    "CacheEfficiency",
    "RagEfficiency",
    "AgentEfficiency",
    "ServiceAttribution",
):
    check(sheet in wb.sheetnames, f"Excel export writes {sheet}")
check("CostSummary" not in wb.sheetnames, "Excel export does not write pricing/cost summary")

wb.close()
shutil.rmtree(out_dir, ignore_errors=True)

print("\nRESULT:", "all checks passed" if _failures == 0 else f"{_failures} FAILURE(S)")
sys.exit(1 if _failures else 0)
